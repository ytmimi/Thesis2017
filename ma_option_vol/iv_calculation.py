import re
import openpyxl
import datetime as dt
from update_excel_workbooks import find_index_0
from wallstreet.blackandscholes import BlackandScholes as BS
from CONSTANTS import (TREASURY_WORKSHEET, TOTAL_TREASURY_SHEET_ROWS, DATE_COLUMN, THREE_MONTH_COLUMN, SIX_MONTH_COLUMN, TWELVE_MONTH_COLUMN,
OPTION_SHEET_PATTERN_INT,OPTION_SHEET_PATTERN_FLOAT,STOCK_SHEET_PATTERN )



def calculate_implied_vol(stock_price, strike_price, days_till_expiration, option_price, risk_free_rate, option_type, dividend_yeild ):
	'''
	Given The stock price, options price, days till expiration, and risk free rate, the implied volatility for the option is calculated

	stock_price				Should be an integer or a float

	strike_price			Should be an integer or a float

	days_till_expiration	Should be an integer

	option_price			Should be an integer or a float

	risk_free_rate			Should be an integer or a float

	option_type				Should be a sting of either 'call' or 'put'

	dividend_yeild 			optional argument. Should either be a integer or a float
	'''
	#instance of the BlackandScholes class from the wallstreet module
	option = BS(S=stock_price, K=strike_price, T=days_till_expiration, price=option_price, r=risk_free_rate, option=option_type, q=dividend_yeild)
	return option.impvol	


def days_till_expiration(start_date, expiration_date):
	'''
	Given an expiration date, and a a starting date, the days to expiration is calculated
	'''
	return (expiration_date-start_date).days


def find_starting_risk_free_rate_index(start_date, data_start_row):
	'''
	Search the TREASURY_SHEET for the start_date provided.

	start_date 		Should be a datetime object
	'''
	date_index= find_index_0(worksheet=TREASURY_WORKSHEET,start=data_start_row, end=TOTAL_TREASURY_SHEET_ROWS,date_col=DATE_COLUMN, date_0=start_date)
	return date_index


def is_negative(num):
	'''
	if num is negative, 0 is returned else, num is returned
	'''
	if num <0:
		return True
	else:
		return False


def calculate_sheet_iv(stock_sheet, option_sheet,sheet_date_column,sheet_price_column,data_start_row,data_end_row, three_month_data_col=4,six_month_data_col=5, twelve_month_data_col=6 , three_month=True, six_month=False, twelve_month=False):
	'''
	Given a stock_sheet, and an option_sheet implied volatility is calculated for each row of the option_sheet that contains price data
	'''
	#get the option_type,expiration_date, and stirke_price from the option_sheet.
	option_type= option_sheet['B3'].value
	expiration_date= option_sheet['B4'].value
	strike_price= option_sheet['B5'].value
	
	#sets the starting index for the TREASURY_WORKSHEET
	starting_rf_index= find_starting_risk_free_rate_index(start_date=option_sheet['B9'].value, data_start_row=2)

	#iterate through each row of the option_sheet
	for (index, i) in enumerate(range(data_start_row, data_end_row+1)):
		#gets the value in the date column
		date= option_sheet.cell(row=i,column=sheet_date_column).value
		if date == None:
			break
		else:
			#get the options price
			option_price = option_sheet.cell(row= i, column= sheet_price_column).value

			if three_month:
				if option_price == 0:
					option_sheet.cell(row=i,column= three_month_data_col).value = 0
				else:
					#calcualte the days till expiration:
					days_per_year= (days_till_expiration(start_date=date, expiration_date=expiration_date)/365)
					#get the stock price from the stock_sheet
					stock_price= stock_sheet.cell(row=i, column= sheet_price_column).value
					#get the risk free rate from the TREASURY_WORKSHEET
					rf= TREASURY_WORKSHEET.cell(row=starting_rf_index+index, column=THREE_MONTH_COLUMN).value
					#if rf is negative, set rf to 0
					if is_negative(rf):
						rf=0
					ivol = calculate_implied_vol(stock_price= stock_price, strike_price= strike_price, days_till_expiration= days_per_year, 
												 option_price=option_price, risk_free_rate=rf, option_type=option_type, dividend_yeild=0)
					#sets the cell in the option_sheet to ivol
					option_sheet.cell(row=i, column=three_month_data_col).value= ivol

			if six_month:
				if option_price == 0:
					option_sheet.cell(row=i,column= six_month_data_col).value = 0
				else:
					#calcualte the days till expiration:
					days_per_year= (days_till_expiration(start_date=date, expiration_date=expiration_date)/365)
					#get the stock price from the stock_sheet
					stock_price= stock_sheet.cell(row=i, column= sheet_price_column).value
					#get the risk free rate from the TREASURY_WORKSHEET
					rf= TREASURY_WORKSHEET.cell(row=starting_rf_index+index, column=SIX_MONTH_COLUMN).value
					#if rf is negative, set rf to 0
					if is_negative(rf):
						rf=0
					ivol = calculate_implied_vol(stock_price= stock_price, strike_price= strike_price, days_till_expiration= days_per_year, 
												 option_price=option_price, risk_free_rate=rf, option_type=option_type, dividend_yeild=0)
					#sets the cell in the option_sheet to ivol
					option_sheet.cell(row=i, column=six_month_data_col).value= ivol

			if twelve_month:
				if option_price == 0:
					option_sheet.cell(row=i,column= twelve_month_data_col).value = 0
				else:
					#calcualte the days till expiration:
					days_per_year= (days_till_expiration(start_date=date, expiration_date=expiration_date)/365)
					#get the stock price from the stock_sheet
					stock_price= stock_sheet.cell(row=i, column= sheet_price_column).value
					#get the risk free rate from the TREASURY_WORKSHEET
					rf= TREASURY_WORKSHEET.cell(row=starting_rf_index+index, column=TWELVE_MONTH_COLUMN).value
					#if rf is negative, set rf to 0
					if is_negative(rf):
						rf=0
					ivol = calculate_implied_vol(stock_price= stock_price, strike_price= strike_price, days_till_expiration= days_per_year, 
												 option_price=option_price, risk_free_rate=rf, option_type=option_type, dividend_yeild=0)
					#sets the cell in the option_sheet to ivol
					option_sheet.cell(row=i, column=twelve_month_data_col).value= ivol


def calculate_workbook_iv(workbook_path, sheet_date_column, sheet_price_column, data_start_row, three_month_data_col=4, six_month_data_col=5, twelve_month_data_col=6, three_month=True, six_month=False, twelve_month=False):
	'''
	Calculates implied volatility for the data in each options worksheet.
	'''
	#loads the given workbook
	wb = openpyxl.load_workbook(workbook_path)

	#iterates through all the sheets in the workbook:
	for index, sheet_name in enumerate(wb.get_sheet_names()):
		#if the sheet_name matches the STOCK_SHEET_PATTERN
		if re.match(STOCK_SHEET_PATTERN, sheet_name):
			#sets the stock sheet
			stock_sheet = wb.get_sheet_by_name(sheet_name)
			#gets the total rows from the stock sheet
			stock_sheet_rows = stock_sheet.max_row

		#if the sheet_name matches either of the OPTION_SHEET_PATTERNS'
		elif re.match(OPTION_SHEET_PATTERN_FLOAT, sheet_name) or re.match(OPTION_SHEET_PATTERN_INT,sheet_name):
			#sets the option sheet
			option_sheet = wb.get_sheet_by_name(sheet_name)
			
			if three_month:
				#set the column header
				option_sheet.cell(row=data_start_row-1, column=three_month_data_col).value = '3 Month IVOL'
				#calculates the implied volatility for each row of the given sheet
				calculate_sheet_iv(stock_sheet=stock_sheet, option_sheet=option_sheet, sheet_date_column=sheet_date_column,sheet_price_column=sheet_price_column,
								three_month_data_col=three_month_data_col,six_month_data_col=six_month_data_col , twelve_month_data_col=twelve_month_data_col, 
								data_start_row=9, data_end_row=stock_sheet_rows, three_month=True, six_month=False, twelve_month=False)

			if six_month:
				#set the column header
				option_sheet.cell(row=data_start_row-1, column=six_month_data_col).value ='6 Month IVOL'
				#calculates the implied volatility for each row of the given sheet
				calculate_sheet_iv(stock_sheet=stock_sheet, option_sheet=option_sheet, sheet_date_column=sheet_date_column,sheet_price_column=sheet_price_column,
								three_month_data_col=three_month_data_col,six_month_data_col=six_month_data_col , twelve_month_data_col=twelve_month_data_col, 
								data_start_row=9, data_end_row=stock_sheet_rows, three_month=False, six_month=True, twelve_month=False)

			if twelve_month:
				#set the column header 
				option_sheet.cell(row=data_start_row-1, column=twelve_month_data_col).value = '12 Month IVOL'
				#calculates the implied volatility for each row of the given sheet
				calculate_sheet_iv(stock_sheet=stock_sheet, option_sheet=option_sheet, sheet_date_column=sheet_date_column,sheet_price_column=sheet_price_column,
								three_month_data_col=three_month_data_col,six_month_data_col=six_month_data_col , twelve_month_data_col=twelve_month_data_col, 
								data_start_row=9, data_end_row=stock_sheet_rows, three_month=False, six_month=False, twelve_month=True)
	#save the workbook:
	wb.save(workbook_path)
	print('Done calculating IVOL. Saving workbook...')
 










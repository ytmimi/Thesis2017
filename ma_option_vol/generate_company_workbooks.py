import openpyxl
import datetime as dt
import os

import add_bloomberg_excel_functions as abxl
from data_workbooks import Merger_Sample_Data 
from update_excel_workbooks import store_data_to_txt_file
from CONSTANTS import ACQUIRER_DIR, TARGET_DIR, MERGER_SAMPLE

class Create_Company_Workbooks():
    def __init__(self, source_sheet_name, source_file=MERGER_SAMPLE, target_path=TARGET_DIR, acquirer_path=ACQUIRER_DIR):
        self.source_sheet_name = source_sheet_name
        self.source_file = source_file
        self.target_path = target_path
        self.acquirer_path = acquirer_path
        self.wb = openpyxl.load_workbook(source_file)
        self.sample_manager = Merger_Sample_Data(self.wb, source_sheet_name)

    def create_company_workbooks(self):
        #iterates over all the rows of the sample worksheet
        for i in range(1, self.sample_manager.ws_length):
            self.new_company_workbook(row_index=i+1, path=self.target_path)
            self.new_company_workbook(row_index=i+1, path=self.acquirer_path)

    def format_source_data(self, row_index, include):
        indexes = [self.sample_manager.headers.index(x)+1 for x in include]
        #get the row data
        data = self.sample_manager.row_values(row_index, include)
        #set some data values
        one_year = dt.timedelta(days=360)
        start_date = ((data[indexes[2]]['value'])-one_year).date()
        announcement_date = self.adjust_to_weekday(data[indexes[2]]['value']).date()
        end_date = data[indexes[-1]]['value'].date()
        #list of data to use in the sheet
        data = [[include[0], data[indexes[0]]['value']], 
                [include[1], data[indexes[1]]['value']],
                ['Type', 'Equity'],
                ['Start Date', start_date],
                ['Announcement Date', announcement_date],
                ['End Date', end_date],
                ['Formated Start Date',int(start_date.strftime('%Y%m%d'))],
                ['Formated Announcement Date',int(announcement_date.strftime('%Y%m%d'))],
                ['Formated End Date',int(end_date.strftime('%Y%m%d'))],]
        return data

    def new_company_workbook(self,row_index, path):
        if path == self.target_path:
            include = ['Target Name', 'Target Ticker', 'Announce Date', 'Completion/Termination Date']
        elif path == self.acquirer_path:
            include = ['Acquirer Name', 'Acquirer Ticker', 'Announce Date', 'Completion/Termination Date']
        data = self.format_source_data(row_index, include)
        cut_off_date = dt.date(2012,2,15)
        #if the start_date is greater than the cut off date, then create a new file
        if data[3][1] > cut_off_date:
            new_wb = openpyxl.Workbook()
            new_sheet = new_wb.active
            new_sheet.title = 'Options Chain'
            for item in data:
                new_sheet.append(item)
            self.get_company_options_tickers(new_sheet, data[3][1], data[4][1], 10, 1, 90, 'B2', 'B3')
            # save_new_workbook(self, wb, file_name, start_date, path='.', extension='xlsx'):
            self.save_new_workbook(new_wb, data[0][1], data[4][1], path=path,)

    def get_company_options_tickers(self,reference_sheet, start_date, announcement_date, row, column, interval, ticker_cell, type_cell):
        #loop through and call the BDS function while to start_date+the interval is less than 1 months past the announcement date
        while start_date < (announcement_date + dt.timedelta(days=30)):
            bds = abxl.add_BDS_OPT_CHAIN(ticker_cell, type_cell, start_date.strftime('%Y%m%d'))
            reference_sheet.cell(row=row,column=column).value = bds
            start_date += dt.timedelta(days=interval)
            column +=2

    def save_new_workbook(self, wb, file_name, start_date, path='.', extension='xlsx'):
        final_path = self.formated_wb_path(file_name, start_date, extension, path=path)
        #checks to see if the given workbook_path exists
        if os.path.exists(path):
            wb.save(final_path)
        else:
            #if the path doesn't exist, create it
            os.makedirs(path, exist_ok=True)
            print(f'Generating file path: {path}')
            wb.save(final_path)

    def formated_wb_path(self, file_name, date, file_extension='xlsx', path='.'):
        '''Returns the formated path used to save the file'''
        date = date.strftime('%Y-%m-%d')
        file_name = file_name.replace('/', '_')
        file_with_extension = f'{file_name}_{date}.{file_extension}'
        return os.path.join(path, file_with_extension)

    def adjust_to_weekday(self, date):
        '''
        A given datetime object is checked to see whether it is a weekend.  
        If it is, the date is adjusted to the next monday.
        '''
        #dt.weekday() returns a number from 0-6 corresponding to Monday-Sunday
        if date.weekday() == 5: 
            #adjusted to Monday
            date += dt.timedelta(days=2) 
        elif date.weekday() == 6:
            #adjusted the date to Monday
            date += dt.timedelta(days=1) 
        return date 




import openpyxl
import os
import re
import datetime as dt
from CONSTANTS import CALL_DESIGNATION_PATTERN, PUT_DESIGNATION_PATTERN


def group_contracts_by_strike(reference_wb):
    '''
    Given a workbook with many option contracs, those that have the same strike price are grouped together.
    Returns a dictionary with keys representing the strike and type of the option (call or put), and the value being a list of
    sheet names corresponding to each option with the given strike.
    '''
    #a dictionary to store the sorted data
    options_contracts = {}

    #exclude the first sheet because that isn't an options contract
    contract_list = reference_wb.get_sheet_names()[1:]

    #loop through each contract sheet:
    for (index, contract) in enumerate(contract_list):
        #split the sheet name by whitespace and take only the last item in the list
        #the last item will either look similar to 'C(some numbers)' or 'P(some numbers)'
        contract_type = contract.split(' ')[-1]
        
        #if the contract is a call set the default value, create a new key for the contract if it
        #doesn't already exist, increase the count by 1, and append the contract to the appropriate list 
        if re.match(CALL_DESIGNATION_PATTERN, contract_type):
            options_contracts.setdefault('call',{})
            options_contracts['call'].setdefault(contract_type, [])
            options_contracts['call'][contract_type].append(contract)

        #if the contract is a put set the default value, create a new key for the contract if it
        #doesn't already exist, increase the count by 1, and append the contract to the appropriate list
        elif re.match(PUT_DESIGNATION_PATTERN, contract_type):
            options_contracts.setdefault('put',{})
            options_contracts['put'].setdefault(contract_type, [])
            options_contracts['put'][contract_type].append(contract)

    #finally return the options_contracts dictionary
    return options_contracts



def group_contracts_by_expiration(reference_wb):
	'''
	Given a workbook with many option contracs, those that have the same expiration date are grouped together.
	Returns a dictionary with keys representing the expiration date and type of the option (call or put), and the value being a list of
	sheet names corresponding to each option with the given expiration date.
	'''
	#a dictionary to store the sorted data
	options_contracts = {}

	#exclude the first sheet because that isn't an options contract
	contract_list = reference_wb.get_sheet_names()[1:]

	#loop through each contract sheet:
	for (index, contract) in enumerate(contract_list):
		#split the sheet name by whitespace and take only the last item in the list
		#the last item will either look similar to 'C(some numbers)' or 'P(some numbers)'
		contract_type = contract.split(' ')[-1]
		expiration = contract.split(' ')[2]

		#if the contract is a call set the default value, create a new key for the contract if it
		#doesn't already exist, increase the count by 1, and append the contract to the appropriate list 
		if re.match(CALL_DESIGNATION_PATTERN, contract_type):
		    options_contracts.setdefault('call',{})
		    options_contracts['call'].setdefault(expiration, [])
		    options_contracts['call'][expiration].append(contract)

		#if the contract is a put set the default value, create a new key for the contract if it
		#doesn't already exist, increase the count by 1, and append the contract to the appropriate list
		elif re.match(PUT_DESIGNATION_PATTERN, contract_type):
		    options_contracts.setdefault('put',{})
		    options_contracts['put'].setdefault(expiration, [])
		    options_contracts['put'][expiration].append(contract)

	#finally return the options_contracts dictionary
	return options_contracts


def create_sorted_sheet(new_workbook, reference_wb, new_sheet_title, reference_sheet_list, data_start_row, data_column, index_column):
    '''
    Given the following arguments, a new worksheet is created in the new workbook, with sorted data taken from worksheets of the reference_wb

    new_workbook          should be a new openpyxl Workbook where the data will be stored
    
    reference_wb          should be an openpyxl.workbook.Workbook containig the sheets contined in reference_sheet_list
    
    
    new_sheet_title       should be a string with the name that you would like to give the newly created sheet
            
    
    reference_sheet_list  should be a list of sheets that you would like to pull data from
    
    data_start_row        should be an integer indicating which row or the reference_sheet the data stars on
    
    data_column           should be a list of numbers, where 1=column A, 2=column B, 3=column C, and so on. 
                          specifies which columns of data are of interest from the reference_sheet contained in
                          in reference_sheet_list
    
    
    index_column          should be a list of numbers, where 1=column A, 2=column B, 3=column C, and so on.
                          specifies which columns from the reference_sheet_list[0], you want to use as your index's
                          NOTE: For accurate indexing, each sheet of data referenced should have the same index
    
    '''
    #creates a new sheet where the combined data will be stored and names it based on the passed in new_sheet_title argument
    #If the only sheet is the default 'Sheet', then rename it to the desired new_sheet_title,
    if new_workbook.get_sheet_names()[0] == 'Sheet':
        new_sheet = new_workbook.get_active_sheet()
        new_sheet.title = new_sheet_title 
    #else create an entirely new sheet and give it the apporpiate new_sheet_title
    else:
         new_sheet = new_workbook.create_sheet(title = new_sheet_title )
        
    
    #formats the data_column, and index_column in case they were imput as letters.
    data_column = convert_to_numbers(data_column)
    index_column = convert_to_numbers(index_column)


    #iterate over all the sheets in the reference_wb that were passed in reference_sheet_list
    for (index,sheet_name) in enumerate(reference_sheet_list):  
        #if its the first iteration, update the sheet with the index_colum and the data_column
        if index ==0:
            #load in the new worksheet.
            data_sheet = reference_wb.get_sheet_by_name(sheet_name)
            
            #get the max rows of the loaded worksheet
            max_row = data_sheet.max_row
            
            #iterate over the columns and the rows that we're interested in
            for (index,column_num) in enumerate(index_column+data_column):
                #find the max_column of the new_worksheet
                max_col = new_sheet.max_column
                
                #if we reach the fist item in the data column set a header for the contract name
                if column_num == data_column[0]:
                    #the header is set at the current max_col+1, which data will go under
                    new_sheet.cell(row= data_start_row-1, column= max_col+1).value = sheet_name
                    
                    
                #iterate over row indexes from data_start_row to max_row+1
                for i in range(data_start_row, max_row+1):
                    #if its the first iteration of adding data to the sheet
                    if index == 0:
                        new_sheet.cell(row=i, column=max_col).value = data_sheet.cell(row= i, column= column_num).value
                    
                    else:
                    	#if the value is a datetime object, then make sure the value of the cell is a formated date
                    	if type(data_sheet.cell(row= i, column= column_num).value) == dt.datetime:
                        	new_sheet.cell(row=i, column=max_col+1).value = data_sheet.cell(row= i, column= column_num).value.date()
                    	else:
                        	new_sheet.cell(row=i, column=max_col+1).value = data_sheet.cell(row= i, column= column_num).value

                                

        #else: just grab the data_column    
        else:
            #load in the new worksheet.
            data_sheet = reference_wb.get_sheet_by_name(sheet_name)
            
            #get the max rows of the loaded worksheet
            max_row = data_sheet.max_row
            
            #iterate over the columns and rows that we're interested in
            for (index,column_num) in enumerate(data_column):
                #find the max_column of the new_worksheet
                max_col = new_sheet.max_column
                
                #if we reach the fist item in the data column set a header for the contract name
                if column_num == data_column[0]:
                    #the header is set at the current max_col+1, which data will go under
                    new_sheet.cell(row= data_start_row-1, column=max_col+1).value = sheet_name
                
                #iterate over row indexes from data_start_row to max_row+1
                for i in range(data_start_row, max_row+1):
                    new_sheet.cell(row=i, column=max_col+1).value = data_sheet.cell(row= i, column= column_num).value


def convert_to_numbers(lst):
    for (index, value) in enumerate(lst):
        if type(value) == str:
            lst[index] = openpyxl.utils.column_index_from_string(value)
    return lst



def create_sorted_workbooks(reference_wb_path, header_start_row, data_column, index_column, sort_by_strike=True, sort_by_expiration=False):
	'''
	Creates a new workbook, containing sorted data in each of its sheets

	reference_wb_path   should be the path to the workbook that data will be taken from

	header_start_row      is a variable that will be passed to the call of create_sorted_sheet()

	data_column         is a variable that will be passed to the call of create_sorted_sheet()

	index_column        is a variable that will be passed to the call of create_sorted_sheet()
	'''
	#initial check to ensure that the user has designated one sort method, but not both. By defualt sort_by_stike is active.
	if(sort_by_strike and sort_by_expiration):
		return print('Error: Choose either to sort by strike or by expiration')

	elif((not sort_by_strike) and (not sort_by_expiration)):
		return print('Error: Choose to sort by strike or by expiration')

	#loads the reference workbook
	reference_wb = openpyxl.load_workbook(reference_wb_path)

	#creats the new workbooks.
	new_call_wb = openpyxl.Workbook()
	new_put_wb = openpyxl.Workbook()

	#adds the stock sheet to the new_call_wb.  This is called outside of the for loop
	create_sorted_sheet(new_workbook =new_call_wb, reference_wb= reference_wb,
						new_sheet_title= 'Stock Price', reference_sheet_list=[reference_wb.get_sheet_names()[1]],
						data_start_row= header_start_row, data_column= [data_column[0]],
						index_column= index_column)

	#adds the stock sheet to the new_put_wb.  This is called outside of the for loop
	create_sorted_sheet(new_workbook =new_put_wb, reference_wb= reference_wb,
						new_sheet_title= 'Stock Price', reference_sheet_list=[reference_wb.get_sheet_names()[1]],
						data_start_row= header_start_row, data_column= [data_column[0]],
						index_column= index_column)

	#sets the value of contracts based on the type of sort designated
	if sort_by_strike:
		#stores the outpuf of group_contracts_by_strike(), which is a nested dictionary that looks similar to:
		#{ {'call':{'C55':['List of contract sheets'], ....}, {'put':{'P55':['List of contract sheets']} } } }
		contracts=group_contracts_by_strike(reference_wb= reference_wb)

	elif sort_by_expiration:
		#stores the outpuf of group_contracts_by_expiration(), which is a nested dictionary that looks similar to:
		#{ {'call':{'03-20-15':['List of contract sheets'], ....}, {'put':{'03-20-15':['List of contract sheets']} } } }
		contracts = group_contracts_by_expiration(reference_wb= reference_wb)

	#iterate over all the keys in the 'call' dictionary stored in contracts 
	for (index, key) in enumerate(contracts['call']):
		#create a sorted sheet in the new workbook for every strike price
		create_sorted_sheet(new_workbook =new_call_wb, reference_wb= reference_wb,
							new_sheet_title= key, reference_sheet_list=contracts['call'][key],
							data_start_row= header_start_row, data_column= data_column,
							index_column= index_column)


	#iterate over all the keys in the 'put' dictionary stored in contracts 
	for (index, key) in enumerate(contracts['put']):
		#create a sorted sheet in the new workbook for every strike price
		create_sorted_sheet(new_workbook =new_put_wb, reference_wb= reference_wb,
							new_sheet_title= key, reference_sheet_list=contracts['put'][key],
							data_start_row= header_start_row, data_column= data_column,
							index_column= index_column)

	if sort_by_strike:
		#save new_call_wb
		save_new_workbook(new_workbook=new_call_wb, workbook_path= reference_wb_path,
						new_folder='call_by_strike', append_file_name='call_by_strike')
		#save new_put_wb
		save_new_workbook(new_workbook= new_put_wb, workbook_path= reference_wb_path,
						new_folder= 'put_by_strike', append_file_name='put_by_strike')

	elif sort_by_expiration:
		#save new_call_wb
		save_new_workbook(new_workbook=new_call_wb, workbook_path= reference_wb_path,
							new_folder='call_by_expiration', append_file_name='call_by_expiration')
		#save new_put_wb
		save_new_workbook(new_workbook= new_put_wb, workbook_path= reference_wb_path,
						new_folder= 'put_by_expiration', append_file_name='put_by_expiration')


def save_new_workbook(new_workbook,workbook_path,new_folder,append_file_name):
    '''
    Saves the new file into a newly created folder at the end of the current path file path, 
    and appends the file name.
    '''
    #breakes the path into a list split by '/'
    path_list = workbook_path.split('/')

    #the last item in the split_path is removed and returned from pop().
    #By defualt pop removes and returns the last item, but it is explicilty removed below.
    #the file name is the last item, and we split that by the '.' yeilding a new list in the form ['file_name','file_extension']
    split_file_name = path_list.pop(-1).split('.')
    
    #creates a new file name from the existing name, and the new_folder's value
    new_file_name = '{}_{}'.format(split_file_name[0], append_file_name)

    #appends the name of the new folder to the end of the path
    path_list.append(new_folder)

    #takes the list and converts it to a string representing the new path.
    new_path ='/'.join(path_list)

    
    #checks to see if the given new_path exists
    if os.path.exists(new_path):
        #formats the new file_path with the variables created above
        final_path = '{}/{}.{}'.format(new_path,new_file_name,split_file_name[-1])
        #save the workbook
        new_workbook.save(final_path)
        print('Saving {}.{}'.format(new_file_name,split_file_name[-1]))    
    else:
        #if the path doesn't exist, create it
        os.makedirs(new_path, exist_ok=False)
        print('Generating file path: {}'.format(new_path))
        final_path = '{}/{}.{}'.format(new_path,new_file_name,split_file_name[-1])
        #save the workbook
        new_workbook.save(final_path)
        print('Saving {}.{}'.format(new_file_name,split_file_name[-1]))
 












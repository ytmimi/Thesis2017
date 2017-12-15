import re
import openpyxl
import sys
import os


#Root path to all the files
ROOT_PATH = os.path.abspath(os.pardir)

#path to acquirer folder
ACQUIRER_DIR = os.path.join(ROOT_PATH,'company_data', 'acquirer')
#path to the sorted call by strike folders in the acquirer folder
ACQUIRER_SCBS_DIR = os.path.join(ACQUIRER_DIR, 'call_by_strike')
#path to the sorted put by strike folders in the acquirer folder
ACQUIRER_SPBS_DIR = os.path.join(ACQUIRER_DIR, 'put_by_strike')
#path to the sorted call by expiration folder in the acquirer folder
ACQUIRER_SCBE_DIR = os.path.join(ACQUIRER_DIR, 'call_by_expiration')
#path to the sorted put by expiration folders in the acquirer folder
ACQUIRER_SPBE_DIR = os.path.join(ACQUIRER_DIR, 'call_by_expiration')
#path to the atm vol folder in the acquirer folder
ACQUIRER_ATM_VOL_DIR = os.path.join(ACQUIRER_DIR, 'atm_vol')

#path to the target folder
TARGET_DIR = os.path.join(ROOT_PATH,'company_data', 'target')
#path to the sorted call by strike folders in the target folder
TARGET_SCBS_DIR = os.path.join(TARGET_DIR, 'call_by_strike')
#path to the sorted put by strike folders in the target folder
TARGET_SPBS_DIR = os.path.join(TARGET_DIR, 'put_by_strike')
#path to the sorted call by expiration folder in the target folder
TARGET_SCBE_DIR = os.path.join(TARGET_DIR, 'call_by_expiration')
#path to the sorted put by expiration folders in the target folder
TARGET_SPBE_DIR = os.path.join(TARGET_DIR, 'call_by_expiration')
#path to the atm vol folder in the target folder
TARGET_ATM_VOL_DIR = os.path.join(TARGET_DIR, 'atm_vol')

#path to the output folder
OUTPUT_DIR = os.path.join(ROOT_PATH, 'company_data', 'function_output')

#path to the sample folder
SAMPLE_DIR = os.path.join(ROOT_PATH, 'company_data', 'sample')

#sample file path
MERGER_SAMPLE = os.path.join(SAMPLE_DIR,'M&A List A-S&P500 T-US Sample Set.xlsx')


#a regular expression for a formated option description where the strike is an integer
OPTION_SHEET_PATTERN_INT = re.compile(r'^\w+\s\w+\s\d{2}-\d{2}-\d{2}\s\w+$')

#a regular expression for a formated option description where the strike is a float
OPTION_SHEET_PATTERN_FLOAT= re.compile(r'^\w+\s\w+\s\d{2}-\d{2}-\d{2}\s\w+\.\w+$')

#a regular expression pattern for the stock sheet
STOCK_SHEET_PATTERN =re.compile(r'^\w+\s\w+\s\w+$')

#a regular expression for a formated option description where the strike is an integer
OPTION_DESCRIPTION_PATTERN_INT= re.compile(r'^\w+\s\w+\s\d{2}/\d{2}/\d{2}\s\w+$')

#a regular expression for a formated option description where the strike is a foat
OPTION_DESCRIPTION_PATTERN_FLOAT = re.compile(r'^\w+\s\w+\s\d{2}/\d{2}/\d{2}\s\w+\.\w+$')

#a regular expression to designate whether an option description is a Call
CALL_DESIGNATION_PATTERN = re.compile(r'[C]\d+')

#a regular expression to designate whether an option description is a Put
PUT_DESIGNATION_PATTERN = re.compile(r'[P]\d+')



#usefull information for the treasury sheet
#TREASURY_WORKBOOK_PATH = '{}/{}/{}/{}'.format(os.path.abspath(os.pardir), 'company_data','sample', 'Week_Day Treasury Rates.xlsx')
#sets the path to the treasury worksheet
TREASURY_WORKBOOK_PATH = os.path.join(SAMPLE_DIR,'Week_Day Treasury Rates.xlsx')
#loads the sheet title 'Rates' from the treasury workbook
TREASURY_WORKSHEET= openpyxl.load_workbook(TREASURY_WORKBOOK_PATH, data_only=True).get_sheet_by_name('Rates')
#gets the total rows of the treasury sheet
TOTAL_TREASURY_SHEET_ROWS = TREASURY_WORKSHEET.max_row
TREASURY_DATA_START_ROW = 2
DATE_COLUMN= 2
THREE_MONTH_COLUMN= 7
SIX_MONTH_COLUMN= 8 
TWELVE_MONTH_COLUMN= 9

#useful information for the vix data sheet
VIX_INDEX_PATH = os.path.join(SAMPLE_DIR,'VIX INDEX.xlsx')
VIX_SHEET = openpyxl.load_workbook(VIX_INDEX_PATH, data_only=True).get_sheet_by_name('VIX Data')
TOTAL_VIX_SHEET_ROWS = VIX_SHEET.max_row
VIX_DATA_START_ROW = 2
VIX_DATE_COLUMN = 4
VIX_PX_LAST_COL = 5


#total number of minutes in a 365 day year
MINUTES_PER_YEAR= 525600
#total number of minutes in a 30 day period
MINUTES_PER_MONTH= 43200





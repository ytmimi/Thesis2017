import openpyxl
import re
import os
from statistics import mean
from update_excel_workbooks import copy_data, find_index_0
from generate_sorted_options_workbooks import mdy_string_to_date, string_to_strike, group_contracts_by_expiration, sort_expiration_dates
from CONSTANTS import (ACQUIRER_DIR, TARGET_DIR, ACQUIRER_ATM_VOL_DIR, TARGET_ATM_VOL_DIR, 
					VIX_SHEET,TOTAL_VIX_SHEET_ROWS,VIX_DATA_START_ROW, VIX_DATE_COLUMN, VIX_PX_LAST_COL)


def create_atm_vol_workbook(reference_wb_path,row_index=9,index_col=1,cut_off_index_val=15, vix_data_col=8, date_col=2, stock_col=3, ivol_3col=5, ivol_6col=6, ivol_12col=7,call=True):
	'''
	Given a reference option_workbook, a new workbook
	'''
	#load the reference workbook
	wb = openpyxl.load_workbook(reference_wb_path)
	#load the stock sheet, the 2nd sheet in the workbook. Index of 1 corresponds with the 2nd sheet
	stock_sheet = wb.get_sheet_by_name(wb.get_sheet_names()[1])
	total_rows = stock_sheet.max_row
	#by default, finds the row of with a value of 15 in the INDEX column
	end_stock_sheet_index = find_cut_off_index_row(stock_sheet,total_rows,index_col,cut_off_index_val)
	
	#create a new workbook:
	new_wb = openpyxl.Workbook()
	#creates the vol sheet for puts and calls and copies stock data to them
	call_vol_series_sheet = create_vol_sheet(new_wb, call=call)
	copy_stock_data_to_vol_sheet(ref_stock_sheet=stock_sheet, vol_sheet=call_vol_series_sheet, start_row_index=row_index, 
								end_row_index=end_stock_sheet_index, index_col=index_col, date_col=date_col, stock_col=stock_col)

	put_vol_series_sheet = create_vol_sheet(new_wb, call=not(call))
	copy_stock_data_to_vol_sheet(ref_stock_sheet=stock_sheet, vol_sheet=put_vol_series_sheet, start_row_index=row_index, 
								end_row_index=end_stock_sheet_index, index_col=index_col, date_col=date_col, stock_col=stock_col)

	#add the atm implied volatility series to each sheet
	add_atm_vol_series(reference_wb=wb, reference_stock_ws=stock_sheet, ivol_ws=call_vol_series_sheet,
						row_index=row_index,index_col=index_col,cut_off_index_val=cut_off_index_val, 
						date_col=date_col, stock_col=stock_col,call=call, 
						ivol_3col=ivol_3col, ivol_6col=ivol_6col, ivol_12col=ivol_12col)

	add_atm_vol_series(reference_wb=wb, reference_stock_ws=stock_sheet, ivol_ws=put_vol_series_sheet,
						row_index=row_index,index_col=index_col,cut_off_index_val=cut_off_index_val, 
						date_col=date_col, stock_col=stock_col,call=not(call), 
						ivol_3col=ivol_3col, ivol_6col=ivol_6col, ivol_12col=ivol_12col)
	
	#add the vix values for each day to the sheet
	add_vix_series(reference_sheet=call_vol_series_sheet, data_start_row=row_index, date_column=date_col, vix_data_column=vix_data_col)
	add_vix_series(reference_sheet=put_vol_series_sheet, data_start_row=row_index, date_column=date_col, vix_data_column=vix_data_col)
	
	#add column headers to each sheet
	add_column_headers(reference_sheet=call_vol_series_sheet, header_row=row_index-1)
	add_column_headers(reference_sheet=put_vol_series_sheet, header_row=row_index-1)

	#save the workbook
	save_atm_vol_series(reference_wb_path=reference_wb_path, new_wb=new_wb)



def copy_stock_data_to_vol_sheet(ref_stock_sheet, vol_sheet, start_row_index, end_row_index, index_col, date_col, stock_col):
	'''
	Given a reference stock sheet, data in the index, date, and stock column are copied over
	'''
	#copy the index column from the stock sheet
	copy_data(ref_stock_sheet, vol_sheet,start_row_index, end_row_index, index_col, index_col)
	#copy the date column from the stock sheet
	copy_data(ref_stock_sheet, vol_sheet,start_row_index, end_row_index, date_col, date_col)
	#copy the stock price column from the stock sheet
	copy_data(ref_stock_sheet, vol_sheet,start_row_index, end_row_index, stock_col, stock_col)


def add_atm_vol_series(reference_wb, reference_stock_ws, ivol_ws,row_index=9,index_col=1,cut_off_index_val=15, date_col=2, stock_col=3, ivol_3col=5, ivol_6col=6, ivol_12col=7,call=True):
	'''
	creates an implied volatility curve using the closest atm options
	'''
	#creates a dictionary that organizes sheets by expiration date
	contracts = group_contracts_by_expiration(reference_wb)
	#sets the key for the dictionary to either 'call' or 'put'
	option_key = set_key(call)
	#uses the key to create a sorted list of the nested dictionary key's
	sorted_list = sorted_put_or_call_list(contracts, option_key,call=call)

	#iterate up to and including index row 15 by default b/c that is the last day in the event window.
	while reference_stock_ws.cell(row=row_index, column=index_col).value <= cut_off_index_val:
		#sets the current date and stock price from the row index
		curr_date = reference_stock_ws.cell(row=row_index, column=date_col).value.date()
		curr_stock_price = reference_stock_ws.cell(row=row_index, column=stock_col).value
		
		#finds the date key from the sorted list
		near_exp = find_nearest_exp_month(curr_date, sorted_list)
		
		#gets the option_list using the the near_exp
		option_list = contracts[option_key][near_exp]
		
		#consolidate the option list to only those strikes that traded:
		traded_option_list = get_traded_strikes(reference_wb= reference_wb, option_sheet_list=option_list, row=row_index)
		
		#gets the option sheet closest to ATM
		near_atm_sheet = find_nearest_atm_strike(curr_stock_price, traded_option_list)
		
		#load the closest to ATM option sheet
		option_sheet = reference_wb.get_sheet_by_name(near_atm_sheet)	
		
		#get the implied volatility from each interest rate column
		vol3= option_sheet.cell(row=row_index, column=ivol_3col).value
		vol6= option_sheet.cell(row=row_index, column=ivol_6col).value
		vol12= option_sheet.cell(row=row_index, column=ivol_12col).value

		#set the value of option_sheet, vol3, vol6, and vol12 to the vol_series_sheet
		ivol_ws.cell(row=row_index, column=4).value = near_atm_sheet
		ivol_ws.cell(row=row_index, column=5).value = vol3
		ivol_ws.cell(row=row_index, column=6).value = vol6
		ivol_ws.cell(row=row_index, column=7).value = vol12

		row_index +=1


def add_vix_series(reference_sheet, data_start_row,date_column, vix_data_column):
	'''
	Given a reference sheet, the corresponding vix data for each date will be added
	'''
	total_rows = reference_sheet.max_row
	first_date = reference_sheet.cell(row=data_start_row, column=date_column).value
	vix_start_index = find_vix_starting_index(first_date)
	vix_end_index = (vix_start_index+(total_rows-data_start_row))
	
	#iterate over each row of the reference_sheet and add the corresponding vix_sheet value
	offset = 0
	for i in range(data_start_row, total_rows+1):
		reference_sheet.cell(row=i, column=vix_data_column).value = VIX_SHEET.cell(row=vix_start_index+offset, column=VIX_PX_LAST_COL).value
		offset+=1


def find_vix_starting_index(date):
	'''
	Given a date, the corresponding row index of the VIX SHEET is returned
	'''
	index =find_index_0(worksheet=VIX_SHEET,start=VIX_DATA_START_ROW, end=TOTAL_VIX_SHEET_ROWS, 
				date_col=VIX_DATE_COLUMN, date_0=date)

	return index

def find_cut_off_index_row(sheet,row_index,index_col,index_value ):
	'''
	Returns the row in the given sheet that corresponds to the given index_value.
	Starts at the end of the sheet and works its way up
	'''
	while sheet.cell(row=row_index, column=index_col).value != index_value:
		row_index -=1

	return row_index


def create_vol_sheet(reference_wb, call_vol_sheet_index=2, put_vol_sheet_index=3,call=True):
	'''
	Given a reference openpyxl Workbook object, a volatility sheet is added
	'''
	#the first sheet in the workbook
	first_sheet = reference_wb.get_sheet_by_name(reference_wb.get_sheet_names()[0])
	#if the title of the first sheet is 'Sheet'
	if first_sheet.title == 'Sheet':
		if call:
			title = 'ATM_Call_vol'
			first_sheet.title = title
			vol_sheet = first_sheet
			
			#if the sheet doesn't already exist, create it
		#else: load the sheet
		else:
			title = 'ATM_Put_vol'
			first_sheet.title = title
			vol_sheet = first_sheet
			
	#if the title isn't sheet, creat a new one
	else:
		if call:
			title = 'ATM_Call_vol'
			if title not in reference_wb.get_sheet_names():
				new_sheet=reference_wb.create_sheet()
				new_sheet.title = title
				vol_sheet= new_sheet
			else:
				vol_sheet = reference_wb.get_sheet_by_name(title)

		else:
			title = 'ATM_Put_vol'
			if title not in reference_wb.get_sheet_names():
				new_sheet=reference_wb.create_sheet()
				new_sheet.title = title
				vol_sheet= new_sheet
			else:
				vol_sheet = reference_wb.get_sheet_by_name(title)
		
	return vol_sheet

def get_traded_strikes(reference_wb, option_sheet_list, row, month3_col=5, month6_col=6, month12_col=7):
	'''
	Given an option sheet, the implied volatily value for the given row is checked.
	if the row value != 0, then the sheet is appended to a consolidted list
	'''	
	consolidated_list= []
	for sheet in option_sheet_list:
		#load the worksheet
		option_sheet = reference_wb.get_sheet_by_name(sheet)
		ivol3 = option_sheet.cell(row=row, column=month3_col).value
		ivol6 = option_sheet.cell(row=row, column=month6_col).value
		ivol12 = option_sheet.cell(row=row, column=month12_col).value
		if (ivol3!=0) and (ivol6!=0) and (ivol12!=0):
			consolidated_list.append(sheet)

	#if consolidated_list is empty just return the originally passed in list
	if consolidated_list == []:
		return option_sheet_list
	else:
		return consolidated_list

def find_nearest_atm_strike(stock_price, option_sheet_list):
	'''
	Given a stock_price and a list of option_sheets in the form ticker, expiration date, type-strike: 
	return the sheet with the smallest ablosute difference between the stock and strike price
	'''
	#loop through each option sheet and create a list of absolute differences between the stock price and stirke
	abs_diff_list=[]
	#loop through all the sheets and append the absolute difference of the strike and stock price
	for sheet in option_sheet_list:
		str_strike = sheet.split(' ')[-1]
		num_strike = string_to_strike(str_strike)
		abs_diff = round(abs(stock_price - num_strike), 4)
		abs_diff_list.append(abs_diff)

	atm_index = abs_diff_list.index(min(abs_diff_list))
	return option_sheet_list[atm_index]


def find_nearest_exp_month(date, date_list, date_limit= 8):
	'''
	Given a list of sorted expiration dates, the nearest expiration with more days than the date limit left to expiration is returned
	'''
	for date_str in date_list:
		exp_date = mdy_string_to_date(date_str)
		if (exp_date -date).days >=8:
			return date_str


def find_next_exp_month(date, date_list):
	'''
	Given a list of sorted expiration dates, the next expiration after the current near term expiration is returned
	'''
	#finds the near expiration date
	near = find_nearest_exp_month(date, date_list)
	#try except just in case the date_list isn't long enough, then next-index would cause an error
	try:
		#Since the date_list is assumed to be ordered, the next-term expiration proceeds the near-term expiration
		next_index = date_list.index(near)+1
		next_exp = date_list[next_index]
	except Exception as e:
		next_exp = 'NONE'
	
	return next_exp

def set_key(call=True):
	'''
	Sets the key for the option dictionary. returnes either 'call' or 'put'
	'''
	if call:
		dict_key = 'call'
	else:
		dict_key = 'put'

	return  dict_key


def sorted_put_or_call_list(option_dict, dict_key, call=True):
	'''
	Given an option dictionary, a sorted list of the dictionary's keys is returned
	'''
	if call:
		#creates a sorted list of expiration dates
		sorted_list = sort_expiration_dates(list(option_dict[dict_key].keys()))
	else:
		#creates a sorted list of expiration dates
		sorted_list = sort_expiration_dates(list(option_dict[dict_key].keys()))

	return sorted_list

def add_column_headers(reference_sheet,header_row, start_col=1,header_list=['INDEX', 'DATA', 'STOCK_PRICE', 'OPTION', '3month_IVOL', '6month_IVOL','12month_IVOL','VIX_DATA']):
	'''
	Given a reference sheet, the header_list is added to the header row for each column:
	'''
	for i,header in enumerate(header_list,start=start_col):
		reference_sheet.cell(row=header_row, column=i).value = header

def save_atm_vol_series(reference_wb_path, new_wb):
	'''
	Given a reference_wb_path and a newly created wb, the file is either saved in the ACQUIRER_ATM_VOL_DIR, or TARGET_ATM_VOL_DIR
	'''
	split_path_list = reference_wb_path.split('/')
	new_wb_name = '{}_{}'.format('atm_vol', split_path_list.pop(-1))
	base_reference_path = '/'.join(split_path_list)

	if base_reference_path == ACQUIRER_DIR:
		final_path= '{}/{}'.format(ACQUIRER_ATM_VOL_DIR, new_wb_name)
		#if the final path exists
		if os.path.exists(ACQUIRER_ATM_VOL_DIR):
			new_wb.save(final_path)
		#the folder doesn't exist
		else:
			os.makedirs(ACQUIRER_ATM_VOL_DIR, exist_ok=False)
			new_wb.save(final_path)

	elif base_reference_path == TARGET_DIR:
		final_path='{}/{}'.format(TARGET_ATM_VOL_DIR, new_wb_name)
		if os.path.exists(TARGET_ATM_VOL_DIR):
			new_wb.save(final_path)
		#the folder doesn't exist
		else:
			os.makedirs(TARGET_ATM_VOL_DIR, exist_ok=False)
			new_wb.save(final_path)
	else:
		print('ISSUE')



def create_average_vol_sheet(reference_wb_path, data_start_row=9):
	'''
	Given a reference workbook with atm put and call implied volatility data, a new sheet is created to store the average data
	'''
	wb = openpyxl.load_workbook(reference_wb_path)
	#load the call and put sheet
	call_sheet = wb .get_sheet_by_name('ATM_Call_vol')
	put_sheet = wb .get_sheet_by_name('ATM_Put_vol')
	#get the total rows: note, the total rows is the same from both the call and put sheet
	total_rows = call_sheet.max_row
	#create the average sheet
	average_sheet = wb.create_sheet()
	average_sheet.title = 'Average vol'
	#copy over the index, date, stock price from the call sheet
	copy_data(call_sheet, average_sheet,index_start_row=data_start_row-1, index_end_row=total_rows, reference_data_column=1, main_data_column=1)
	copy_data(call_sheet, average_sheet,index_start_row=data_start_row-1, index_end_row=total_rows, reference_data_column=2, main_data_column=average_sheet.max_column+1)
	copy_data(call_sheet, average_sheet,index_start_row=data_start_row-1, index_end_row=total_rows, reference_data_column=3, main_data_column=average_sheet.max_column+1)
	#copy over the options used from the call sheet
	copy_data(call_sheet, average_sheet,index_start_row=data_start_row-1, index_end_row=total_rows, reference_data_column=4, main_data_column=average_sheet.max_column+1)
	#copy over the options used from the put sheet
	copy_data(put_sheet, average_sheet,index_start_row=data_start_row-1, index_end_row=total_rows, reference_data_column=4, main_data_column=average_sheet.max_column+1)
	#get the average of all implied volatility columns
	average_near_put_call_atm_vol(sheet1=call_sheet, sheet2=put_sheet, average_sheet=average_sheet, data_end_row=total_rows, data_start_row=data_start_row)
	#copy over the VIX data from the call sheet
	copy_data(call_sheet, average_sheet,index_start_row=data_start_row-1, index_end_row=total_rows, reference_data_column=8, main_data_column=average_sheet.max_column+1)
	
	#add column header
	add_column_headers(reference_sheet=average_sheet,header_row=data_start_row-1, start_col=4,header_list=['Call', 'Put', '3month_IVOL', '6month_IVOL','12month_IVOL'])

	#replace zero's with cell references:
	replace_zero_with_cell_ref(reference_wb=wb, reference_sheet= average_sheet)

	#calculate the implied volatility change using excel's ln function
	average_sheet_IVC_calculation(reference_wb=wb,reference_sheet= average_sheet)

	#add new column headers
	add_column_headers(reference_sheet=average_sheet, header_row=data_start_row-1, start_col=11, header_list=['IVC_3m', 'IVC_6m', 'IVC_12m', 'VIX_change'])

	#save the workbook
	wb.save(reference_wb_path)

	
def average_near_put_call_atm_vol(sheet1, sheet2, average_sheet, data_end_row, data_start_row=9, month3_col=5, month6_col=6, month12_col=7):
	'''
	takes the average of the values in sheet1 and sheet2 and stores them in the average sheet
	'''
	for i in range(data_start_row, data_end_row+1):
		#gets all implied volatility calculations from sheet1
		s1_month3_vol = sheet1.cell(row=i, column=month3_col).value
		s1_month6_vol = sheet1.cell(row=i, column=month6_col).value
		s1_month12_vol= sheet1.cell(row=i, column=month12_col).value

		#gets all implied volatility calculations from sheet2
		s2_month3_vol = sheet2.cell(row=i, column=month3_col).value
		s2_month6_vol = sheet2.cell(row=i, column=month6_col).value
		s2_month12_vol= sheet2.cell(row=i, column=month12_col).value

		#sets the average implied volatility calculation in the average sheet
		average_sheet.cell(row=i, column=month3_col+1).value = vol_average(s1_month3_vol, s2_month3_vol)
		average_sheet.cell(row=i, column=month6_col+1).value = vol_average(s1_month6_vol, s2_month6_vol)
		average_sheet.cell(row=i, column=month12_col+1).value= vol_average(s1_month12_vol, s2_month12_vol)


def vol_average(num1, num2):
	'''
	computes the average of two numbers
	'''
	if num1 == num2 == 0:
		average = 0
	
	elif num1 == 0:
		average = num2
	
	elif num2 == 0:
		average = num1
	
	else:
		average = (num1+num2)/2

	return average


def replace_zero_with_cell_ref(reference_wb, reference_sheet, data_start_row=9, column_list=[6,7,8]):
	'''
	Go through every row designated. if the value of the cell is zero, the value will be replaced with a cell reference to the cell above it
	'''
	total_rows = reference_sheet.max_row

	#iterate over each column:
	for index, column_num in enumerate(column_list):
		#iterate over each row:
		#import pdb; pdb.set_trace()
		for i in range(data_start_row, total_rows+1):
			#if the cell has a value of 0 set its value to the cell above it
			curr_cell_val = reference_sheet.cell(row=i, column=column_num).value
			prev_cell_val = reference_sheet.cell(row=i-1, column=column_num).value
			#check if the curret cell has a value of zero
			if  curr_cell_val == 0:
				#check to make sure that the value of the previous cell is a number or a cell reference to another cell
				if is_a_number(prev_cell_val) or is_a_cell_reference(prev_cell_val):
					prev_cell_coordinate = reference_sheet.cell(row=i-1, column=column_num).coordinate
					reference_sheet.cell(row=i, column=column_num).value = '={}'.format(prev_cell_coordinate)
				#else just set the value to zero
				else:
					reference_sheet.cell(row=i, column=column_num).value = 0


def is_a_number(input):
	'''
	checks if the value is a number
	'''
	try:
		float(input)
		return True
	except ValueError:
		return False

def is_a_cell_reference(input_):
	'''
	checks to see if the input is a cell reference
	'''
	cell_ref = re.compile(r'=\w+')
	if re.match(cell_ref, input_):
		return True
	else:
		return False

def average_sheet_IVC_calculation(reference_wb,reference_sheet,data_start_row=9, column_list=[6,7,8,9], offset=5 ):
	'''
	Given a workbook path and a reference_sheet, the IVC is calculated using excel's LN function for each row of the workbook
	'''
	total_rows = reference_sheet.max_row

	#iterate over each column
	for index, column_num in enumerate(column_list):
		#iterate over each row:
		for i in range(data_start_row, total_rows+1):
			if i > data_start_row:
				curr_cell_coordinate = reference_sheet.cell(row=i, column=column_num).coordinate
				curr_cell_value = reference_sheet.cell(row=i, column=column_num).value
				prev_cell_coordinate = reference_sheet.cell(row=i-1, column=column_num).coordinate
				prev_cell_value = reference_sheet.cell(row=i-1, column=column_num).value
				
				#check to make sure the value of the previous cell wasn't zero, else you'll get a divide by zero error
				#check to make sure the value of the current cell isn't zero, else you'll get an #num error b/c ln(0) is undefined
				if prev_cell_value != 0 and curr_cell_value !=0:
					reference_sheet.cell(row=i, column=column_num+offset).value = excel_ln_calculation(curr_cell_coordinate, prev_cell_coordinate)
				else:
					#if it is zero, just set the new cell value to zero
					reference_sheet.cell(row=i, column=column_num+offset).value = 0


def excel_ln_calculation(curr_cell, prev_cell):
	'''
	Given the coordinates of the current and previous cell a string for the ln function is returned
	'''
	return '=LN({}/{})'.format(curr_cell, prev_cell)



def add_mean_and_market_model(reference_wb_path):
	'''
	Given a reference path to an atm_vol workbook, the mean model, and market model sheets are added
	'''
	wb = openpyxl.load_workbook(reference_wb_path, data_only= True)

	create_mean_model_sheet(wb)
	create_market_model_sheet(wb)

	wb.save(reference_wb_path)

def create_mean_model_sheet(reference_wb, reference_sheet='Average vol', title='Mean Model', data_start_row=9, 
	start_col=5,data_columns=[1,11,12,13],
	header_list=['AIVC_3m','AIVC_6m', 'AIVC_12m',' ',  'CAIVC_3m','CAIVC_6m', 'CAIVC_12m']):
	'''
	Given a reference workbook, the mean model sheet is added completely formated with averages calculated
	'''
	reference_wb
	
	average_sheet = reference_wb.get_sheet_by_name(reference_sheet)
	total_rows = average_sheet.max_row

	model_sheet = reference_wb.create_sheet(title=title)

	#iterate over all data columns and copy them over to the model_sheet
	for i, column_num in enumerate(data_columns, start=1):
		copy_data(average_sheet, model_sheet, data_start_row-1, total_rows, column_num, i)

	add_column_headers(model_sheet,header_row=data_start_row-1, start_col=start_col,header_list=header_list)
	
	cut_off_row=find_cut_off_index_row(average_sheet,total_rows,index_col=1,index_value=-16)
	
	add_column_headers(model_sheet,header_row=2, start_col=2,header_list=['3m_mean', '6m_mean', '12m_mean'])
	add_excel_average(model_sheet, data_start_row, cut_off_row, data_column=2, cell_output1='B3',cell_output2='B4')
	add_excel_average(model_sheet, data_start_row, cut_off_row, data_column=3, cell_output1='C3',cell_output2='C4')
	add_excel_average(model_sheet, data_start_row, cut_off_row, data_column=4, cell_output1='D3',cell_output2='D4')



def add_excel_average(reference_wb, data_start_row, data_end_row, data_column=2, cell_output1='B3', cell_output2='B4'):
	'''
	Goes through all the cells of a workbook from the start row to the end row, and if the cell value !=0 it is added to the average
	the output is the excel AVERAGE() function placed in the cell_output
	'''
	#list to store non-zero cells
	non_zero_cell =[]
	cell_values = []
	#loop through all the rows of a designated data_column
	for i in range(data_start_row,data_end_row+1):
		cell_value =reference_wb.cell(row=i, column=data_column).value
		#if the value of a cell isn't zero
		if ((cell_value !=0) and (cell_value !=None)) :
			cell_coordinate = reference_wb.cell(row=i, column=data_column).coordinate
			non_zero_cell.append(cell_coordinate)
			cell_values.append(cell_value)

	list_to_str = ','.join(non_zero_cell)
	reference_wb[cell_output1] = '=AVERAGE({})'.format(list_to_str)
	reference_wb[cell_output2] = mean(cell_values)


def create_market_model_sheet(reference_wb, reference_sheet='Average vol', title='Market Model', data_start_row=9, 
	start_col=6,data_columns=[1,11,12,13,14],
	header_list=['E(IVC)_3m','E(IVC)_6m', 'E(IVC)_12m',' ','AIVC_3m','AIVC_6m', 'AIVC_6m',' ',  'CAIVC_3m','CAIVC_6m', 'CAIVC_12m']):
	'''
	Given a reference workbook, the market model sheet is added completely formated
	'''
	average_sheet = reference_wb.get_sheet_by_name(reference_sheet)
	total_rows = average_sheet.max_row

	model_sheet = reference_wb.create_sheet(title=title)

	#iterate over all data columns and copy them over to the model_sheet
	for i, column_num in enumerate(data_columns, start=1):
		copy_data(average_sheet, model_sheet, data_start_row-1, total_rows, column_num, i)

	add_column_headers(model_sheet,header_row=data_start_row-1, start_col=start_col,header_list=header_list)

	model_sheet['B1'] = '3 month'
	model_sheet['C1'] = '6 month'
	model_sheet['D1'] = '12 month'
	model_sheet['A2'] = 'Intercept'
	model_sheet['A3'] = 'Slope'
	model_sheet['A4'] = 'R^2'
	model_sheet['A5'] = 'Standard Error'
	










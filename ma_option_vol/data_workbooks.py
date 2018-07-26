import openpyxl
import datetime as dt
import re
from statistics import mean, stdev
from math import ceil, floor, log
from CONSTANTS import (MERGER_SAMPLE, TREASURY_WORKBOOK_PATH, VIX_INDEX_PATH)
from CONSTANTS import OPTION_DESCRIPTION_PATTERN_INT as odpi
from CONSTANTS import OPTION_DESCRIPTION_PATTERN_FLOAT as odpf
from CONSTANTS import OPTION_SHEET_PATTERN_INT as ospi 
from CONSTANTS import OPTION_SHEET_PATTERN_FLOAT as ospf 
		
from add_bloomberg_excel_functions import Bloomberg_Excel as bbe
from options import Option

def is_row_idx_in_range(func):
	def inner(*args, **kwarg):
		if kwarg['row'] <= args[0].ws_length:
			return func(*args, **kwarg)
		else:
			raise IndexError(
				f'''The sheet has {args[0].ws_length} rows. Unable to get data from row {kwarg['row']}''')
	return inner

def is_col_idx_in_range(func):
	def inner(*args, **kwarg):
		if kwarg['col'] <= args[0].ws_width:
			return func(*args, **kwarg)
		else:
			raise IndexError(
				f'''The sheet has {args[0].ws_width} columns. Unable to get data from column {kwarg['col']}''')
	return inner

class Data_WorkSheet:
	def __init__(self, workbook, sheet_name, header_index=1):
		self.wb = workbook
		self.ws = self.wb[sheet_name]
		self.header_index = header_index

	def __str__(self):
		return self.ws.title

	@property
	def ws_length(self):
		return self.ws.max_row

	@property
	def ws_width(self):
		return self.ws.max_column

	@property	
	def headers(self, row=None):
		head = []
		for i in range(1, self.ws_width+1):
			head.append(self.ws.cell(row=self.header_index, column=i).value)
		return head

	@is_row_idx_in_range
	def row_values(self, *,row):
		'''
		Given a row_index, all values are returned in a dictionary
		in the form {indx:{'column':[index number], 'value':[value]}...} 
		'''
		values = {}
		for i, col in enumerate(self.headers, start=1):
			value = self.ws.cell(row=row, column=i).value
			values[i] = {'column':col, 'value':value}
		return values

	@is_row_idx_in_range
	@is_col_idx_in_range
	def get_value(self, *, row, col):
		return self.ws.cell(row=row, column=col).value

	def get_coordinate(self, row, col):
		''' Given a column and a row, an excel coordinate is returned'''
		return self.ws.cell(row=row, column=col).coordinate

	def set_value(self, row, col, value):
		self.ws.cell(row=row, column=col).value = value

	def row_index_by_date(self, date, date_col=2):
		'''
		assumes that the dates in the worksheet are sorted
		returns the row index of a date value in a sheet if it exists
		else an indexError is raised
		'''
		start_idx = self.header_index+1
		end_idx = self.ws_length
		idx_diff =  end_idx - start_idx + 1
		count = 0
		found = False
		while not found:
			if count <log(idx_diff, 2):
				count +=1
			else:
				raise IndexError('Date not found')
			avg_idx = self.check_index_bounds(floor(mean([start_idx, end_idx])))
			curr_date = self.get_value(row=avg_idx, col=date_col)
			# print(f'{avg_idx}: {curr_date}, {date} Sheet: {self}')
			if (date == curr_date):
				found = True
			elif (date > curr_date):
				start_idx = avg_idx +1
			elif (date < curr_date):
				end_idx = avg_idx -1
		return avg_idx

	def check_index_bounds(self, idx):
		if idx <= self.header_index+1: return self.header_index+1
		elif idx >=self.ws_length: return self.ws_length
		else: return idx

	def copy_data(self, ref_sheet, ref_col, main_col, 
				ref_start_row=None, ref_end_row=None):
		'''
		Duplicates column data in the reference sheet into this sheet
		'''
		if ref_start_row == None and ref_end_row == None:
			for i in range(ref_sheet.header_index, ref_sheet.ws_length+1):
				value = ref_sheet.get_value(row=i, col=ref_col)
				if value != None:
					self.set_value(i, main_col, value)
				else:
					break
		else:
			for i in range(ref_start_row, ref_end_row+1):
				value = ref_sheet.get_value(row=i, col=ref_col)
				if value != None:
					self.set_value(i, main_col, value)
				else:
					continue

	def letter_to_col_index(self, *letters):
		'''returns a list of column indexes A=1, B=2, C=3, etc'''
		return [openpyxl.utils.column_index_from_string(x) for x in letters]

class Merger_Sample_Data(Data_WorkSheet):
	def __init__(self, workbook=openpyxl.load_workbook(MERGER_SAMPLE), 
				sheet_name='Filtered Sample Set'):
		super().__init__(workbook, sheet_name)

	def row_values(self, row, include=[]):
		'''
		Returns a dictionary of key value pairs for the cell values at the given index
		include: a list of column numbers to return in the response
		'''
		data = super().row_values(row=row)
		values = [item['value'] for item in data.values()]
		if len(include) > 0:
			output = {key:value for key, value in data.items() if value['column'] in include}
			return output
		return data

def found_index(func):
	def inner(*args, **kwargs):
		index = args[0].date_indexes.get(kwargs['date'])
		if index == None:
			index = args[0].row_index_by_date(kwargs['date'])
			args[0].date_indexes[kwargs['date']] = index
		return func(*args, date=kwargs['date'])
	return inner

class Treasury_Sample_Data(Data_WorkSheet):
	def __init__(self, workbook=openpyxl.load_workbook(TREASURY_WORKBOOK_PATH, data_only=True), 
				sheet_name='Rates', header_index=1, date_col=2, col_3m=3, col_6m=4, col_12m=5):
		super().__init__(workbook, sheet_name, header_index)
		self.date_col = date_col
		self.col_3m = col_3m
		self.col_6m = col_6m
		self.col_12m = col_12m
		self.date_indexes = {}

	def row_index_by_date(self, date):
		return super().row_index_by_date(date, date_col=self.date_col)

	@found_index
	def rf_3m_on(self, *, date):
		'''return the 3 month risk free rate on the given date'''
		index = self.date_indexes.get(date)
		value = self.get_value(row=index, col=self.col_3m)/100
		return self.is_negative(value)

	@found_index
	def rf_6m_on(self, *, date):
		'''return the 6 month risk free rate on the given date'''
		index = self.date_indexes.get(date)
		value = self.get_value(row=index, col=self.col_6m)/100
		return self.is_negative(value)

	@found_index
	def rf_12m_on(self, *, date):
		'''return the 12 month risk free rate on the given date'''
		index = self.date_indexes.get(date)
		value = self.get_value(row=index, col=self.col_12m)/100
		return self.is_negative(value)

	@staticmethod
	def is_negative(num):
		if num <0: return 0
		else: return num

class VIX_Sample_Data(Data_WorkSheet):
	def __init__(self, workbook=openpyxl.load_workbook(VIX_INDEX_PATH), sheet_name='VIX Data', 
				header_index=1, date_col=4, px_col=5):
		super().__init__(workbook, sheet_name, header_index)
		self.date_col = date_col
		self.px_col = px_col
		self.date_indexes = {}

	def row_index_by_date(self, date):
		return super().row_index_by_date(date, date_col=self.date_col)

	@found_index
	def get_vix_on(self, *, date):
		'''returns the value of the vix on the given date'''
		index = self.date_indexes.get(date)
		return self.get_value(row=index, col=self.px_col)

class Option_Chain_Sheet(Data_WorkSheet):
	def __init__(self, workbook, sheet_name='Options Chain'):
		super().__init__(workbook, sheet_name,)
		self.company_name = self.ws['B1'].value
		self.ticker = self.ws['B2'].value
		self.type = self.ws['B3'].value
		self.start_date = self.ws['B4'].value
		self.announce_date = self.ws['B5'].value
		self.end_date = self.ws['B6'].value

	def option_BDP_description(self, row, col):
		'''given the coordinates of a cell containing an option ticker,
		the bloomberg BDP function is added to the adjacent cell'''
		bdp = bbe().BDP(self.get_coordinate(row, col), "SECURITY_DES")
		self.set_value(row, col+1, bdp)
		
	def sheet_BDP_description(self, start_row=10, start_col=1):
		unique_ticker = []
		for i in range(start_col, self.ws_width+1, 2):
			for j in range(start_row, self.ws_length+1):
				value = self.get_value(row=j, col=i)
				if  value != None:
					if value not in unique_ticker:
						unique_ticker.append(value)
						self.option_BDP_description(j, i)
				else:
					break

	def is_option_exp_in_range(self, exp_date, from_start=8, past_announcemt=60):
		'''
		return True if the given expeiration date > a certain number of days from the start
		and < a given number of days after the announcement
		'''
		exp_from_start = (exp_date - self.start_date).days
		exp_past_announcemt = (exp_date - self.announce_date).days
		return (exp_from_start > from_start) and (exp_past_announcemt < past_announcemt)

class Stock_Sheet(Data_WorkSheet):
	def __init__(self, workbook, sheet_name, header_index=8, date_col=2, px_col=3):
		super().__init__(workbook, sheet_name, header_index)
		self.date_col = date_col
		self.px_col = px_col
		self.company_name = self.ws['B1'].value
		self.ticker = self.ws['B2'].value
		self.start_date = self.ws['B3'].value
		self.announce_date = self.ws['B4'].value
		self.end_date = self.ws['B5'].value
		self.date_indexes = {}


	def row_index_by_date(self, date):
		return super().row_index_by_date(date, date_col=self.date_col)

	@found_index
	def get_price_on(self, *, date):
		index = self.date_indexes.get(date)
		return self.get_value(row=index, col=self.px_col)

	@found_index
	def get_index_on(self, *, date):
		index = self.date_indexes.get(date)
		return self.get_value(row=index, col=1)

	def px_last_lst(self, px_col=3, start_row=None, end_row=None):
		'''returns the time series of traded prices'''
		if start_row == None and end_row == None:
			indx_range = range(self.header_index+1, self.ws_length+1)
			return [self.get_value(row=i, col=px_col) for i in indx_range if self.get_value(row=i, col=px_col)!=0]
		else:
			indx_range = range(start_row, end_row+1)
			return [self.get_value(row=i, col=px_col) for i in indx_range if self.get_value(row=i, col=px_col) !=0]

	def merger_mean(self):
		start_idx = self.row_index_by_date(self.announce_date)
		lst = self.px_last_lst(start_row=start_idx, end_row=self.ws_length)
		return floor(mean(lst))

	def merger_std(self):
		start_idx = self.row_index_by_date(self.announce_date)
		lst = self.px_last_lst(start_row=start_idx, end_row=self.ws_length)
		return ceil(stdev(lst))

	def historic_mean(self):
		end_idx = self.row_index_by_date(self.announce_date)
		lst = self.px_last_lst(start_row=self.header_index+1, end_row=end_idx)
		return floor(mean(lst))

	def historic_std(self):
		end_idx = self.row_index_by_date(self.announce_date)
		lst = self.px_last_lst(start_row=self.header_index+1, end_row=end_idx)
		return ceil(stdev(lst))

	def is_strike_in_range(self, strike, mm, ms, hm, hs, std_multiple=1.5):
		''' 
		checks that the given strike is either between the historic or merger mean
		std_multiple increase (decrease) the std to check
		'''
		mhigh, mlow = mm+ms*std_multiple, mm-ms*std_multiple
		hhigh, hlow = hm+hs*std_multiple, hm-hs*std_multiple
		return((mlow <= strike <= mhigh) or (hlow <= strike <= hhigh))
		
	def add_index(self):
		''' Fills in the index value of the given sheet'''
		index_0 = self.row_index_by_date(self.announce_date)
		for i in range(self.header_index+1, self.ws_length+1):
			self.set_value(i, 1, i - index_0)


class Option_Sheet(Data_WorkSheet):
	def __init__(self, workbook, sheet_name, header_index=8, date_col=2, px_col=3):
		super().__init__(workbook, sheet_name, header_index)
		self.date_col = date_col
		self.px_col = px_col
		self.ticker = self.ws['B1'].value
		self.description = self.ws['B2'].value
		self.option = Option.from_description(self.description)

	def row_index_by_date(self, date):
		return super().row_index_by_date(date, date_col=self.date_col)
	
	def fill_empty_cells(self, fill_value=0):
		for i in range(3, self.ws_width+1):
			for j in range(self.header_index+1, self.ws_length+1):
				if self.get_value(row=j, col=i) == None:
					self.set_value(j, i, fill_value)

	def get_stock_index(self, stock_sheet, date):
		''' returns the stock index on the given date, or 'N/A'
			if the date is not in the stock_sheet'''
		try:
			return stock_sheet.get_index_on(date=date)
		except IndexError as e:
			return 'N/A'

	def copy_index(self, stock_sheet):
		'''copies the numeric index from the given stock_sheet'''
		for i in range(self.header_index+1, self.ws_length+1):
			date = self.get_value(row=i, col=self.date_col)
			if date != None:
				index = self.get_stock_index(stock_sheet, date)
				self.set_value(i, 1, index)
			else: break

	def add_iv_calculation(self, row, col, date, stock_price, rf_rate):
		option_price = self.get_value(row=row, col=self.px_col)
		if option_price == None or option_price == 0 or stock_price == 0: 
			self.set_value(row, col, 0)
			return 0
		else: 
			iv = self.option.implied_volatility(date, stock_price, option_price, rf_rate)
			self.set_value(row, col, iv)
			return iv

	def add_vega_calculation(self, row, col, date, stock_price, rf_rate):
		option_price = self.get_value(row=row, col=self.px_col)
		if option_price == None or option_price == 0 or stock_price == 0: 
			self.set_value(row, col, 0)
			return 0
		else: 
			vega = self.option.vega(date, stock_price, option_price, rf_rate)
			self.set_value(row, col, vega)
			return vega

	def get_stock_price(self, stock_sheet, date):
		''' Returns the stock price on that date, or zero if the data does not exist'''
		try:
			return stock_sheet.get_price_on(date=date)
		except IndexError as e:
			return 0

	def get_risk_free_rate(self, treasury_sheet, date, rate):
		'''
		Returns the appropriate risk free rate based on the rate argument
		or zero if the data does not exist
		'''
		try:
			if rate != 3 and rate != 6 and rate != 12:
				raise ValueError('Rate must be set to either 3, 6, or 12')
			else:
				if rate == 3: return treasury_sheet.rf_3m_on(date=date)
				elif rate == 6: return treasury_sheet.rf_6m_on(date=date)
				else: return treasury_sheet.rf_12m_on(date=date)
		except IndexError as e:
			return 0

	def sheet_iv_calculation(self, col, stock_sheet, treasury_sheet, *,rate, heading):
		''' 
		col: the colom to store the calculation in 
		stock_sheet: instance of the Stock_Sheet class associated with the option
		teasury_sheet: sheet containing treasury data
		rate: either 3, 6, or 12 depending on which type of treasury you want
		heading: column heading to add to the worksheet
		'''
		self.set_value(self.header_index, col, heading)
		for i in range(self.header_index+1, self.ws_length+1):
			date = self.get_value(row=i, col=self.date_col)
			index = self.get_value(row=i, col=1)
			if date == None:
				break
			elif index == 'N/A':
				self.set_value(i, col, 0)
			else:
				stock_price = self.get_stock_price(stock_sheet, date)
				rf_rate = self.get_risk_free_rate(treasury_sheet, date, rate)
				self.add_iv_calculation(i, col, date, stock_price, rf_rate)

	def sheet_vega_calculation(self,col, stock_sheet, treasury_sheet, *, rate, heading):
		''' 
		col: the colom to store the calculation in 
		stock_sheet: instance of the Stock_Sheet class associated with the option
		teasury_sheet: sheet containing treasury data
		rate: either 3, 6, or 12 depending on which type of treasury you want
		heading: column heading to add to the worksheet
		'''
		self.set_value(self.header_index, col, heading)
		for i in range(self.header_index+1, self.ws_length+1):
			date = self.get_value(row=i, col=self.date_col)
			index = self.get_value(row=i, col=1)
			if date == None:
				break
			elif index == 'N/A':
				self.set_value(i, col, 0)
			else:
				stock_price = self.get_stock_price(stock_sheet, date)
				rf_rate = self.get_risk_free_rate(treasury_sheet, date, rate)
				self.add_vega_calculation(i, col, date, stock_price, rf_rate)

def has_stock_sheet(func):
	'''Decorator that checks if the Option_Workbook has a stock sheet'''
	def inner(*args, **kwargs):
		#self is the first argument aka arg[0]
		if args[0].stock_sheet != None:
			return func(*args, **kwargs)
		else:
			raise AttributeError('Please run create_stock_sheet method')
	return inner

def has_option_sheets(func):
	'''Decorator that checks if the Option_Workbook has option sheets'''
	def inner(*args, **kwargs):
		#self is the first argument aka arg[0]
		if args[0].option_sheetnames != None:
			return func(*args, **kwargs)
		else:
			raise AttributeError('Please run add_option_sheets method')
	return inner


class Option_Workbook:
	def __init__(self, path):
		self.path = path
		self.wb = openpyxl.load_workbook(path)
		self.chain_sheet = Option_Chain_Sheet(self.wb, 'Options Chain')
		self.data_index = ['Index', 'Date']
		self.data_headers = ['PX_LAST']

	@property
	def stock_sheet(self):
		try:
			sheetname = f'{self.chain_sheet.ticker} {self.chain_sheet.type}'
			return Stock_Sheet(self.wb, sheetname)
		except KeyError:
			return None

	@property 
	def option_sheetnames(self):
		''' 
		returns a sorted list of option_sheetnames, or None
		if option sheets have not been created
		'''
		lst = [s for s in self.wb.sheetnames if(re.match(ospi, s) or re.match(ospf, s))]
		if lst: return lst
		else: return None

	def save(self, path=None):
		if path != None:
			self.wb.save(path)
		else:
			self.wb.save(self.path)

	def stock_meta_data(self):
		return [['Company Name', self.chain_sheet.company_name],
			['Ticker', f'{self.chain_sheet.ticker} {self.chain_sheet.type}'],
			['Start Date', self.chain_sheet.start_date],
			['Announcement Date', self.chain_sheet.announce_date],
			['End Date', self.chain_sheet.end_date],]

	def create_stock_sheet(self, BDH_args=None, BDH_vals=None):
		new_sheet = self.wb.create_sheet(index=1)
		new_sheet.title = f'{self.chain_sheet.ticker} {self.chain_sheet.type}'
		for data in self.stock_meta_data():
			new_sheet.append(data)
		for i, header in enumerate(self.data_index + self.data_headers, start=1):
			new_sheet.cell(row=8, column=i).value = header
		new_sheet['B9'].value = bbe().BDH(self.chain_sheet.ticker, self.data_headers[0],
										self.chain_sheet.start_date, self.chain_sheet.end_date,
										type_='Equity', opt_args=BDH_args, opt_vals=BDH_vals)

	def option_meta_data(self, ticker, description):
		option_data = Option.parse_option_description(description)
		data_labels = ['Ticker', 'Description', 'Type', 'Expiration Date', 'Strike Price']
		return zip(data_labels, [ticker, description]+option_data)

	#add a function that adds new option sheets
	def create_option_sheet(self, ticker, description,BDH_args=None, BDH_vals=None):
		new_sheet = self.wb.create_sheet()
		new_sheet.title = description.replace('/', '-')
		for data in self.option_meta_data(ticker, description):
			new_sheet.append(data)
		for i, header in enumerate(self.data_index + self.data_headers, start=1):
			new_sheet.cell(row=8, column=i).value = header
			new_sheet['B9'].value = bbe().BDH(ticker, self.data_headers,self.chain_sheet.start_date,
												'B4',opt_args=BDH_args,opt_vals=BDH_vals)

	@has_stock_sheet
	def add_option_sheets(self, BDH_args=None, BDH_vals=None):
		ss = self.stock_sheet
		mm, ms, = ss.merger_mean(), ss.merger_std()
		hm, hs  = ss.historic_mean(), ss.historic_std()
		for i in range(1, self.chain_sheet.ws_width,2):
			for j in range(10, self.chain_sheet.ws_length+1):
				ticker = self.chain_sheet.get_value(row=j, col=i)
				description = self.chain_sheet.get_value(row=j, col=i+1)
				#check if the ticker and description values are not None
				if self.proper_desciption_format(description) and ticker != None:
					type_, exp_date, strike = Option.parse_option_description(description)
					#bottle neck is the is_strike_in_range()
					if ss.is_strike_in_range(strike, mm, ms, hm, hs) and self.chain_sheet.is_option_exp_in_range(exp_date):
						self.create_option_sheet(ticker, description, BDH_args, BDH_vals)
				elif ticker == None:
					break

	def proper_desciption_format(self, description):
		'''Returns true if the description matches the designated option re'''
		if description == None:
			return False
		else:
			return (re.match(odpi, description) or re.match(odpf, description))

	@has_option_sheets
	def fill_option_sheets(self):
		for sheet in self.option_sheetnames:
			Option_Sheet(self.wb, sheet).fill_empty_cells()

	@has_option_sheets
	@has_stock_sheet
	def add_index_to_sheets(self):
		self.stock_sheet.add_index()
		for sheet in options:
			Option_Sheet(self.wb, sheet).copy_index()

	@has_option_sheets
	@has_stock_sheet
	def calculate_workbook_iv(self, iv_col, rate, heading):
		'''
		iv_col: column to store the results
		rate: the tresury rate to use: 3, 6, or 12
		heading: heading for the vega column added to each sheet of the workbook
		'''
		stock_sheet = self.stock_sheet
		treasury_sheet = Treasury_Sample_Data()
		for sheet in self.option_sheetnames:
			option_sheet = Option_Sheet(self.wb, sheet)
			for i in range(option_sheet.header_index+1, option_sheet.ws_length+1):
				date = option_sheet.get_value(row=i, col=option_sheet.date_col)
				if date != None:
					option_sheet.sheet_iv_calculation(iv_col, stock_sheet, treasury_sheet, rate=rate, heading=heading)
				else:
					break

	@has_option_sheets
	@has_stock_sheet
	def calculate_workbook_vega(self, vega_col, rate, heading):
		'''
		vega_col: column to store the results in
		rate: the tresury rate to use: 3, 6, or 12
		heading: heading for the vega column added to each sheet of the workbook
		'''
		stock_sheet = self.stock_sheet
		treasury_sheet = Treasury_Sample_Data()
		for sheet in self.option_sheetnames:
			option_sheet = Option_Sheet(self.wb, sheet)
			for i in range(option_sheet.header_index+1, option_sheet.ws_length+1):
				date = option_sheet.get_value(row=i, col=option_sheet.date_col)
				if date != None:
					option_sheet.sheet_vega_calculation(vega_col, stock_sheet, treasury_sheet, rate=rate, heading=heading)
				else:
					break

	def delete_sheets(self, include=[]):
		pass







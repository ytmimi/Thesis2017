import openpyxl
import os
import datetime as dt
import re
from statistics import mean, stdev
from math import ceil, floor
import add_bloomberg_excel_functions as abxl
from CONSTANTS import ( OPTION_DESCRIPTION_PATTERN_INT, OPTION_DESCRIPTION_PATTERN_FLOAT, OPTION_SHEET_PATTERN_INT, OPTION_SHEET_PATTERN_FLOAT,
                        STOCK_SHEET_PATTERN, OUTPUT_DIR)

#meant to use with the 'Option Chain' sheet
#########################Tested
def update_sheet_with_BDP_description(workbook_path, sheet_name, starting_col, starting_row):
    '''
    Given an excel workbook, The BDP function is added with the appropriate cell reference and description  
    Note: The excel workbook needs to be opened and saved so that bloomberg data can populate
    '''
    #opens the workbook
    wb = openpyxl.load_workbook(workbook_path)
    #gets the specified sheet from the workbook
    sheet = wb[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    
    #list to keep track of tickers that have already been used
    unique_ticker = []
    #iterate over each column:
    for i in range(starting_col, total_columns+1, 2):
        #iterate over each row:
        for j in range(starting_row, total_rows+1):
            #check if the current cell is blank
            if sheet.cell(row=j, column=i).value == None:
                break
            else:
                #check to see if the cell value is unique
                if sheet.cell(row=j, column=i).value not in unique_ticker:
                    #add the value to the ticker list
                    unique_ticker.append(sheet.cell(row=j, column=i).value)
                    #set the value of the adjacent cell
                    sheet.cell(row=j, column= i+1).value = abxl.add_BDP_fuction(sheet.cell(row=j, column=i).coordinate, "SECURITY_DES")
    
    # #saves the workbook
    wb.save(workbook_path)
    wb_name = workbook_path.split('/')[-1]
    data = '{} contracts sampled for {}\n'.format(len(unique_ticker), wb_name)
    store_data_to_txt_file(file_name='option_des', data=data)

#########################Tested
def update_option_contract_sheets(workbook_path, sheet_name,starting_col,starting_row, sheet_start_date_cell, sheet_announce_date_cell, sheet_end_date_cell,  data_header_row, data_table_index, data_table_header, BDH_optional_arg=None, BDH_optional_val=None):
    '''
    Creates new sheets in the given excel workbook based on Option data stored in the given sheet_name.

    workbook_path       the full file_path to the specified workbook

    sheet_name          Specify the sheet_name that data will be referenced from

    sheet_start_date_cell Give the coordinates of the cell in 

    sheet_end_date_cell Specify the coordinates of the the cell in the specified sheet that contains the end date
    '''
    
    #combine data_table_index and data_table_header
    total_data_headers = data_table_index+data_table_header


    #data labels to be added to the new excel worksheet
    option_data_labels = ['Security Name', 'Description', 'Type', 'Expiration Date', 'Strike Price']
    
    #given the file path, an excel workbook is loaded.
    wb = openpyxl.load_workbook(workbook_path)
    
    #The sheet we want to get data from is set to the variable data_sheet
    data_sheet = wb[sheet_name]
    
    #The cell in the sheet that contains the start date, as passed in by the function.
    if type(data_sheet[sheet_end_date_cell].value) == int:
        start_date = dt.datetime.strptime(str(data_sheet[sheet_start_date_cell].value),'%Y%m%d').date()
    else:
        start_date= data_sheet[sheet_start_date_cell].value.date()

    #The cell in the sheet that contains the announcement date, as passed in by the function.
    if type(data_sheet[sheet_announce_date_cell].value) == int:
        announcement_date = dt.datetime.strptime(str(data_sheet[sheet_announce_date_cell].value),'%Y%m%d').date()
    else:
        announcement_date= data_sheet[sheet_announce_date_cell].value.date()

    total_rows = data_sheet.max_row
    total_columns = data_sheet.max_column

    #counter to keep track of each sheet created
    sheet_count = 0
    #gets the average stock price and standard deviation of the stock price data for the historic and merger period:
    historic = historic_stock_mean_and_std(reference_wb_path=workbook_path, price_column_header='PX_LAST', header_start_row=data_header_row, date_0=dt.datetime.strptime(str(data_sheet[sheet_announce_date_cell].value),'%Y%m%d'))
    merger = merger_stock_mean_and_std(reference_wb_path=workbook_path, price_column_header='PX_LAST', header_start_row=data_header_row, date_0=dt.datetime.strptime(str(data_sheet[sheet_announce_date_cell].value),'%Y%m%d'))

    #iterate through the columns of the data_sheet
    for i in range(starting_col, total_columns+1, 2):

        #iterate through the rows of the data_sheet
        for j in range(starting_row, total_rows+1):
            ticker_cell = data_sheet.cell(row=j, column=i).value
            des_cell = data_sheet.cell(row=j, column=i+1).value
            #check if the ticker and description cell does not = None
            if ((ticker_cell != None) and (des_cell != None)):
                #checks to see if the description cell value follows the pattern of an option description
                if (re.match(OPTION_DESCRIPTION_PATTERN_INT, des_cell) or re.match(OPTION_DESCRIPTION_PATTERN_FLOAT, des_cell)) :

                    #format_option_description() returns the following list:
                    #[security_name, option_description, option_type, expiration_date, strike_price]
                    option_data = format_option_description(ticker_cell, des_cell)

                    #the number of days between the expiration and start date. 
                    expiration_from_start = (option_data[3] - start_date).days
                    #the number of days past the annoucement date and the 
                    days_past_announcemt = (option_data[3]- announcement_date).days

                    #import pdb; pdb.set_trace()
                    #if the expiration date is less than 8 days after the start date or if the expiration date is 60 days paste the announcment date, do nothing.
                    if (expiration_from_start < 8) or (days_past_announcemt > 60) :
                        pass
                        #otherwise, keep creating sheets
                    else:
                        #check to see if the stike is within 1.5 standard deviation of the historical and merger stock mean
                        if ((is_in_range(num=option_data[-1], high=historic[0]+1.5*historic[1], low=historic[0]-1.5*historic[1])) or (is_in_range(num=option_data[-1], high=merger[0]+1.5*merger[1], low=merger[0]-1.5*merger[1]))):
                            #creates a new sheet for the passed in workbook
                            new_sheet = wb.create_sheet()
                            #increment the sheet count by 1
                            sheet_count +=1
                            #'/' aren't allowed in excel sheet names, so we replace them with '-' if the name contains '/' 
                            new_sheet.title = option_data[1].replace('/', '-')

                            #zip creates a tuple pair for each item of the passed in lists. this tuple can then be appended to the sheet
                            for data in zip(option_data_labels,option_data):
                                new_sheet.append(data)

                            #loop through every value of total_data_headers and add it to the worksheet at the specified data_header_row
                            for (index, value) in enumerate(total_data_headers, start= 1) :
                                new_sheet.cell(row = data_header_row,column = index ).value = value 

                            #add the BDH formula to the sheet
                            new_sheet.cell(row=data_header_row+1, column=2).value = abxl.add_option_BDH( security_name = option_data[0],
                                                                                          fields = data_table_header, 
                                                                                          start_date = data_sheet[sheet_start_date_cell].value,
                                                                                          end_date = 'B4',
                                                                                          optional_arg = BDH_optional_arg,
                                                                                          optional_val = BDH_optional_val)
                else:
                    print('Not a valid option description. Could not create new workbook sheets for {}'.format(des_cell))
                    continue
        
    #save the workbook
    wb.save(workbook_path)
    wb_name = workbook_path.split('/')[-1] 
    data='Saving workbook with {} new tabs: {} \n'.format(sheet_count,wb_name)
    store_data_to_txt_file(file_name='option_sheets', data=data)
 
#########################Tested
def format_option_description(security_name, option_description):
    '''
    security_name should be a string that looks similar to 'BBG00673J6L5 Equity'
    option_description should be a string that looks similar to 'PFE US 12/20/14 P18'
    return formatted option data
    '''
    #will split the option_description by whitespace into a list that looks like: ['PFE', 'US', '12/20/14', 'P18']
    description_list = option_description.split(' ')

    #determins the option type based on description_list[-1][0] = 'P'
    if description_list[-1][0] =='P':
        option_type = 'Put'
    elif description_list[-1][0] == 'C':
        option_type = 'Call'

    #description_list[2] = 12/20/14 and convertis it into a datetime object
    expiration_date = dt.datetime.strptime(description_list[2],'%m/%d/%y').date()

    #description_list[-1][1:] = '18', and converts the string to an int
    try:
        strike_price = int(description_list[-1][1:])
    #if the string was a floating point number like 18.5, convert it to a float
    except:
        strike_price = float(description_list[-1][1:])

    option_data_list = [security_name, option_description, option_type, expiration_date, strike_price]

    return option_data_list


def update_workbook_data_index(workbook_path, data_start_row, index_column):
    '''
    Given a workbook, loop through all the sheets of that workbook and update the index for each sheet.
    '''
    #loads an excel workbook given the file path to that workbook.
    wb = openpyxl.load_workbook(workbook_path)
    #gets a list of all the sheets in the workbook
    sheet_list = wb.get_sheet_names()

    #in case index column was passed in as a character, convert it to an integer
    index_column= convert_to_numbers(index_column)

    #iterates through every sheet
    for (index, sheet_name) in enumerate(sheet_list):
        #indexing starts at 0.
        if index == 0:
            #get the announcement date from the first sheet
            sheet = wb.get_sheet_by_name(sheet_name)
            announcement_date = sheet['B5'].value
        #if the sheet_name matches the stock sheet pattern:
        if re.match(STOCK_SHEET_PATTERN, sheet_name):
            #load the stock sheet and save it to the stock_sheet variable
            stock_sheet = wb.get_sheet_by_name(sheet_name)
            total_rows = stock_sheet.max_row
            update_sheet_index(reference_sheet= stock_sheet, date=announcement_date, start_row= data_start_row)

        #elif the sheet_name matches an options contract sheet 
        elif(re.match(OPTION_SHEET_PATTERN_INT, sheet_name) or re.match(OPTION_SHEET_PATTERN_FLOAT, sheet_name)):
            #load the option sheet and save it to the option_sheet variable
            option_sheet = wb.get_sheet_by_name(sheet_name)
            copy_data(reference_sheet=stock_sheet, main_sheet=option_sheet, index_start_row=data_start_row, 
                      index_end_row=total_rows, reference_data_column=index_column, main_data_column=index_column)
    wb.save(workbook_path)
    print('Indexed each sheet. Saving workbook...')


def update_sheet_index(reference_sheet, date, start_row):
    '''
    Given an excel worksheet,a designated date, and a starting row,
    an index is added for each date relative to the specified date and row
    '''
    #gets the total number of rows in the worksheet
    total_rows = reference_sheet.max_row
    
    #returns the row index of the reference_sheet containg the date value
    index_0 =find_index_0(worksheet=reference_sheet,start= start_row, end=total_rows, date_col=2, date_0= date)
    #iterates over every column in the given date_column from the start to the end and add the index value to the cell
    for index in range(start_row, total_rows+1):
        reference_sheet.cell(row= index, column=1).value = index - index_0


###################Will not test
def update_read_data_only(file_path):
    '''
    Opens an Excel workbook in read_only mode, removing links to function calls, but maintaing the values stored in each cell.
    '''
    wb = openpyxl.load_workbook(file_path, data_only= True)
    wb.save(file_path)
    return wb


def store_data_to_txt_file(file_name, data,file_path=OUTPUT_DIR):
    '''
    Given a file path, output data from a function is stored
    '''
    #full file path
    complete_path = '{}/{}.{}'.format(file_path,file_name,'txt')
    #check if the file exists
    if os.path.exists(file_path):
        #if the file exisist open it to append
        f = open(complete_path, 'a')
        f.write(data)
        f.close()
    
    #else creat the file_path
    else:
        os.makedirs(file_path, exist_ok=False)
        f = open(complete_path, 'w')
        f.write(data)
        f.close()
        
#########################Tested
def delet_workbook_option_sheets(workbook_path):
    '''
    Given the file path to a workbook, all the option sheets are deleted
    '''
    wb = openpyxl.load_workbook(workbook_path)
    start_sheet_num = len(wb.sheetnames)
    #set the active sheet in the workbook to the first sheet:
    wb.active = 0
    for (index,sheet) in enumerate(wb.sheetnames):
        #if the sheet is an option sheet
        if(re.match(OPTION_SHEET_PATTERN_INT, sheet)) or (re.match(OPTION_SHEET_PATTERN_FLOAT, sheet)):
            del wb[sheet]

    end_sheet_num = len(wb.sheetnames)
    deleted_sheet_num = start_sheet_num - end_sheet_num 
    wb_name = workbook_path.split('/')[-1]
    data ='Deleted {} sheets from {} \n'.format(deleted_sheet_num, wb_name)
    
    store_data_to_txt_file(file_name= 'deleted_sheets', data= data)
    wb.save(workbook_path)


#########################Tested
def find_index_0(worksheet,start, end, date_col, date_0):
    '''
    binary search function to determine which row index of the worksheet
    contains the date we're looking for.

    worksheet   Should be an openpyxl worksheet object

    start       Should be an index >=1

    end         Should be an index <= total rows of the given worksheet

    date_col    Should be the column containing dates to search through. 1=A, 2=B, 3=C, etc.

    date_0      The specific date to search for
    '''
    #list comprehesion  for all the row indexes.
    index_list = [x for x in range(start,end+1)]
    start_index = index_list[0]
    end_index = index_list[-1]
    average_index = floor((end_index + start_index)/2)
    #variable for the while loop
    found = False
    while not found:
        #print(start_index, found)        
        curr_date = worksheet.cell(row=average_index, column=date_col).value
        if (date_0 == curr_date):
            found = True

        elif (date_0 > curr_date):
            start_index = average_index +1
            average_index = floor((end_index + start_index)/2)

        elif (date_0 < curr_date):
            end_index = average_index -1
            average_index = floor((end_index + start_index)/2)

    return average_index

#########################Tested
def copy_data(reference_sheet, main_sheet,index_start_row, index_end_row, reference_data_column, main_data_column):
    '''
    Copies data from the reference_sheet to the main_sheet.  index_start_row is assumed to be the same for both the reference_sheet and main_sheet

    reference_sheet:        Should be an openpyxl worksheet object that data will be coppied over from.

    main_sheet:             Should be an openpyxl worksheet object taht data will be coppied to.

    index_start_row:        Should be an integer that specifies the row of the worksheet to start copying from.

    index_end_row:          Should be an interger that specifies the last row of the worksheet that data should be coppied from. 

    reference_data_column:  Can either be an integer that specifies which column from the reference worksheet contains the data to copied
                            or the letter associated with the data column 1=A, 2=B, C=3, etc.

    main_data_column:       Can either be an integer that specifies which column in the main worksheet the data should be copied to
                            or the letter associated with the data column 1=A, 2=B, C=3, etc.
    '''
    for i in range(index_start_row, index_end_row+1):
        #if the value is a datetime.datetime object
        if type(reference_sheet.cell(row= i, column= reference_data_column).value) == dt.datetime:
            main_sheet.cell(row=i, column=main_data_column).value = reference_sheet.cell(row=i, column=reference_data_column).value.date()
        elif reference_sheet.cell(row= i, column= reference_data_column).value == None:
            continue
        else:
            main_sheet.cell(row=i, column=main_data_column).value = reference_sheet.cell(row=i, column=reference_data_column).value



def update_stock_price_sheet(workbook_path, sheet_name, stock_sheet_index, sheet_start_date_cell,sheet_announce_date_cell, sheet_end_date_cell,  data_header_row, data_table_index, data_table_header, BDH_optional_arg=None, BDH_optional_val=None ):
    '''
    Adds a sheet with stock price information to the workbook
    '''
    #load the given workbook
    wb = openpyxl.load_workbook(workbook_path)

    #gets the reference sheet
    reference_sheet = wb.get_sheet_by_name(sheet_name)
    ticker = '{} {}'.format(reference_sheet['B2'].value, reference_sheet['B3'].value)

    #create a new sheet, and makes it the second sheet in the workbook. sheet indexing starts at 0.
    new_sheet = wb.create_sheet(index=stock_sheet_index)
    #sets the title of the new worksheet
    new_sheet.title = ticker
    #basic data to be added to the sheet
    data = [('Company Name', reference_sheet['B1'].value),
            ('Company Ticker',ticker),
            ('Start Date', reference_sheet[sheet_start_date_cell].value),
            ('Announcement Date',reference_sheet[sheet_announce_date_cell].value),
            ('End Date',reference_sheet[sheet_end_date_cell].value)]

    #appends the data to the top of the spreadsheet
    for (index,data_lst) in enumerate(data):
        new_sheet.append(data_lst)

    #combines both passed lists:
    total_headers = data_table_index + data_table_header
    #set the index and column headers for the worksheet
    for (index, value) in enumerate(total_headers, start= 1):
        new_sheet.cell(row=data_header_row,column=index).value = value
        if value.upper() == ('DATE'):
            #sets the BDH function into place
            new_sheet.cell(row= data_header_row+1, column= index).value = abxl.add_option_BDH(security_name = data[1][1],
                                                                            fields = data_table_header, 
                                                                            start_date = reference_sheet[sheet_start_date_cell].value,
                                                                            end_date = reference_sheet[sheet_end_date_cell].value,
                                                                            optional_arg = BDH_optional_arg,
                                                                            optional_val = BDH_optional_val)

    #saves the newly added sheet to the workbook.
    wb.save(workbook_path)
    wb_name = workbook_path.split('/')[-1]
    data = 'Added {} sheet to workbook: {}\n'.format(ticker, wb_name)
    store_data_to_txt_file(file_name= 'stock_sheets', data= data)



def update_workbook_average_column(reference_wb_path, column_header, header_row, data_start_row, ignore_sheet_list=[]):
    '''
    Given the path to an excel workbook, Averages are calculated for each sheet of data
    '''
    #loads an excel workbook from the given file_path
    reference_wb = openpyxl.load_workbook(reference_wb_path, data_only=True)
    #returns a dictionary of 'sheet_names':[column data indexes] for each sheet of the given workbook
    sheet_data_columns =find_column_index_by_header(reference_wb= reference_wb, column_header= column_header, header_row= header_row)

    #removes any sheets that are ment to be ignored if provided
    if ignore_sheet_list != []:
        #iterates over every sheet name passed into ignore_sheet_list
        for index, ignore_sheet in enumerate(ignore_sheet_list):
            #removes the sheet name from the dictionary sheet_data_columns, so that it wont be iterated over next
            sheet_data_columns.pop(ignore_sheet)

    #iterate over each key(sheet_name) in sheet_data_columns:
    for (index,key) in enumerate(sheet_data_columns):
        #update the given sheet with the average column
        update_sheet_average_column(reference_wb= reference_wb, 
                                    sheet_name= key,
                                    data_columns= sheet_data_columns[key],
                                    data_start_row= data_start_row,
                                    column_header= column_header)

    #saves the excel workbook
    reference_wb.save(reference_wb_path)
    print('Saving Workbook...')


def update_sheet_average_column(reference_wb,sheet_name,data_columns, data_start_row, column_header):
    '''
    Given an excel worksheet, and a specified list of columns, averages are calcualted for each row of the data
    '''
    #loads the sheet of the reference_wb
    sheet = reference_wb.get_sheet_by_name(sheet_name)

    #gets the max row of the sheet 
    max_row = sheet.max_row

    #gets the max column of the sheet
    max_col = sheet.max_column
    
    #sets the header for the average column to the average_col_header and places it one row above the data
    sheet.cell(row=data_start_row-1, column=max_col+1).value = '{} {} {}'.format(sheet_name, 'Average', column_header)

    #iterate over each row of the workbook:
    for i in range(data_start_row,max_row+1):
        #an empty lest to store the values for the cells of the given row
        cell_values = []
        #iterate over each cell in the data column
        for (index, column_ref) in enumerate(data_columns):
            #if the value of the cell isn't 0, append it to the cell_values list
            if sheet.cell(row=i, column=column_ref).value != 0:
                cell_values.append(sheet.cell(row=i, column=column_ref).value)

        #assing the value of the average column
        #if the cell_values list is an empyt list
        if cell_values == []:
            #set the value of the cell to 0
            sheet.cell(row=i, column=max_col+1).value = 0
        #else cell_values is populated
        else:
            #set the value of the average column to the mean of the cell_values
            sheet.cell(row=i, column=max_col+1).value = statistics.mean(cell_values)


def find_column_index_by_header(reference_wb, column_header, header_row):
    '''
    Returns a dictionary where the key is the sheet name, and the value is the column where the specified header was located
    '''
    #an empty dictionary that will store the sheet_name as the key, and a list of data_columns as the key's value 
    data_columns_by_sheet= {}

    #iterate over all the sheetnames in the workbook
    for (index,sheet_name) in enumerate(reference_wb.sheetnames):
        #load the given worksheet.
        sheet = reference_wb[sheet_name]

        #get the max_column for the worksheet:
        max_col =sheet.max_column

        #add a key in the dictionary for the given sheet
        data_columns_by_sheet.setdefault(sheet_name, [])

        #loop through all the cells in the header_row
        for i in range(max_col+1):
            #If the value in the column header matches the header_value we're searching for, then append the column index to the key's list:
            if  column_header == sheet.cell(row=header_row, column=i+1).value:
                data_columns_by_sheet[sheet_name].append(i+1)

        #if no columns were found, remove that key from the dictionary
        if data_columns_by_sheet[sheet_name] == []:
            data_columns_by_sheet.pop(sheet_name)

    #return the dictionary with the data for each sheet
    return data_columns_by_sheet


def stock_data_to_list(reference_wb,price_column_header, header_start_row, start_index, end_index):
    '''
    Given the file path to a workbook, data in a particular cell is added to a list and then the list is returned
    '''
    #returns a dictionary with {'sheet_name':[data_column]}
    data_column = find_column_index_by_header(reference_wb = reference_wb, column_header= price_column_header, header_row= header_start_row)

    #data list to store all the values:
    data_list = []
    
    #iterate over all the keys in the data_column:
    for (index,key) in enumerate(data_column):
        if re.match(STOCK_SHEET_PATTERN, key):
            #load the worksheet
            sheet=reference_wb[key]
            for i in range(start_index, end_index+1):
                if sheet.cell(row=i,column=data_column[key][0]).value !=0:
                    data_list.append(sheet.cell(row=i,column=data_column[key][0]).value)
    #return the data_list
    return data_list

def data_average(data_list):
    '''
    returns the average of a given list, rounded down to the nearest whole number
    '''
    return floor(mean(data_list))


def data_standard_dev(data_list):
    '''
    returns the standard deviation of a given list, rounded up to the nearest whole number
    '''
    return ceil(stdev(data=data_list))


def historic_stock_mean_and_std(reference_wb_path,price_column_header, header_start_row, date_0):
    '''
    calculates the mean and standard deviation for prices up to the announcemnt date
    '''
    #loads the workbook and the specified sheet
    wb = openpyxl.load_workbook(reference_wb_path)
    #get the second sheet in the workbook
    sheet = wb[wb.sheetnames[1]]

    total_rows=sheet.max_row

    index0=find_index_0(worksheet=sheet,start=header_start_row+1, end=total_rows, date_col=2, date_0=date_0)
    data_list=stock_data_to_list(reference_wb=wb, price_column_header=price_column_header, 
                                 header_start_row=header_start_row, start_index=header_start_row+1, end_index=index0)

    average = data_average(data_list)
    st_dev = data_standard_dev(data_list)

    return(average, st_dev)


def merger_stock_mean_and_std(reference_wb_path, price_column_header, header_start_row, date_0):
    '''
    calculates the mean and standard deviation for prices from the merger date to the end of the M&A
    '''
    wb = openpyxl.load_workbook(reference_wb_path)
    #get the second sheet in the workbook
    sheet = wb[wb.sheetnames[1]]

    total_rows=sheet.max_row

    index0=find_index_0(worksheet=sheet,start=header_start_row+1, end=total_rows, date_col=2, date_0=date_0)
    data_list=stock_data_to_list(reference_wb=wb, price_column_header=price_column_header, 
                                 header_start_row=header_start_row, start_index=index0, end_index=total_rows)

    average = data_average(data_list)
    st_dev = data_standard_dev(data_list)

    return(average, st_dev)


#########################Tested
def is_in_range(num, high, low):
    '''
    Given a number, and a high and low range, True is returned if the number is within the range 
    '''        
    return low <=num <= high


#########################Tested
def fill_option_wb_empty_cells(reference_wb_path, column_start, row_start, fill_value):
    '''
    Goes through each sheet and fills in the blanks with the designated fill_vale
    '''
    #load the workbook
    wb = openpyxl.load_workbook(reference_wb_path)

    #iterate over each sheet
    for (index,sheet_name) in enumerate(wb.sheetnames):
        #if the sheet is an option sheet
        if re.match(OPTION_SHEET_PATTERN_INT, sheet_name) or re.match(OPTION_SHEET_PATTERN_FLOAT, sheet_name):
            sheet = wb[sheet_name]
            fill_option_sheet_empty_cells(reference_sheet=sheet,column_start= column_start, row_start= row_start, fill_value= fill_value)
    
    #save the workbook:
    wb.save(reference_wb_path)
    print('Done filling empty cells with {}.'.format(fill_value))



#########################Tested
def fill_option_sheet_empty_cells(reference_sheet, column_start, row_start, fill_value):
    '''
    Goes through a sheet and fills in the empty cells with the designated fill_value
    '''
    #get the max_rows
    total_rows=reference_sheet.max_row
    #get the max_columns
    total_columns = reference_sheet.max_column
    #iterate_over_each column:
    for i in range(column_start, total_columns+1):
        #iterate over each row:
        for j in range(row_start, total_rows+1):
            if reference_sheet.cell(row=j, column=i).value == None:
                reference_sheet.cell(row=j, column=i).value = fill_value


#########################Tested
def convert_to_numbers(lst):
    '''
    Takes a list or a single character, and returns an integer, where A=1, B=2, C=3,...etc.
    '''
    #if lst is passed in as just a single value
    if type(lst) == str:
        lst = openpyxl.utils.column_index_from_string(lst)
    else:
        #if lst is passed in as a list 
        for (index, value) in enumerate(lst):
            if type(value) == str:
                lst[index] = openpyxl.utils.column_index_from_string(value)
    return lst


def add_extra_sheets(reference_wb_path, sheet_name, ticker_column, description_column,sheet_start_date_cell, sheet_announce_date_cell, sheet_end_date_cell,  data_header_row, data_table_index, data_table_header, BDH_optional_arg=None, BDH_optional_val=None):
    '''
    Given a workbook containing option contract tickers and desctiptions, new sheets are added to the workbook if they don't already exist
    '''
    #combine data_table_index and data_table_header
    total_data_headers = data_table_index+data_table_header
    #data labels to be added to the new excel worksheet
    option_data_labels = ['Security Name', 'Description', 'Type', 'Expiration Date', 'Strike Price']
    #given the file path, an excel workbook is loaded.
    wb = openpyxl.load_workbook(reference_wb_path)
    #The sheet we want to get data from is set to the variable data_sheet
    data_sheet = wb.get_sheet_by_name(sheet_name)
    #gets the total rows of the worksheet
    total_rows = data_sheet.max_row
    #counter to keep track of each sheet created
    sheet_count = 0
    #gets the average stock price and standard deviation of the stock price data for the historic and merger period:
    historic = historic_stock_mean_and_std(reference_wb_path=reference_wb_path, price_column_header='PX_LAST', header_start_row=data_header_row, date_0=dt.datetime.strptime(str(data_sheet[sheet_announce_date_cell].value),'%Y%m%d'))
    merger = merger_stock_mean_and_std(reference_wb_path=reference_wb_path, price_column_header='PX_LAST', header_start_row=data_header_row, date_0=dt.datetime.strptime(str(data_sheet[sheet_announce_date_cell].value),'%Y%m%d'))

    while (data_sheet.cell(row=total_rows, column= description_column).value).replace('/','-') not in wb.get_sheet_names():
        #format_option_description() returns the following list:
        #[security_name, option_description, option_type, expiration_date, strike_price]
        option_data = format_option_description(data_sheet.cell(row=total_rows, column=ticker_column).value,
                                                        data_sheet.cell(row=total_rows, column=description_column).value)

        if (re.match(OPTION_DESCRIPTION_PATTERN_INT, option_data[1]) or re.match(OPTION_DESCRIPTION_PATTERN_FLOAT,option_data[1])):
           
            #check to see if the stike is within 1.5 standard deviation of the historical and merger stock mean
            if ((is_in_range(num=option_data[-1], high=historic[0]+1.5*historic[1], low=historic[0]-1.5*historic[1])) or (is_in_range(num=option_data[-1], high=merger[0]+1.5*merger[1], low=merger[0]-1.5*merger[1]))):
                #creates a new sheet for the passed in workbook
                new_sheet = wb.create_sheet()
                #increment the sheet count by 1
                sheet_count +=1
                #/' aren't allowed in excel sheet names, so we replace them with '-' if the name contains '/' 
                new_sheet.title = option_data[1].replace('/', '-')
                #zip creates a tuple pair for each item of the passed in lists. this tuple can then be appended to the sheet
                for data in zip(option_data_labels,option_data):
                    new_sheet.append(data)
                #loop through every value of total_data_headers and add it to the worksheet at the specified data_header_row
                for (index, value) in enumerate(total_data_headers, start= 1) :
                    new_sheet.cell(row = data_header_row,column = index ).value = value 
                #add the BDH formula to the sheet
                new_sheet['B{}'.format(data_header_row+1)] = abxl.add_option_BDH( security_name = option_data[0],
                                                                              fields = data_table_header, 
                                                                              start_date = data_sheet[sheet_start_date_cell].value,
                                                                              end_date = 'B4',
                                                                              optional_arg = BDH_optional_arg,
                                                                              optional_val = BDH_optional_val)
        total_rows -= 1
    wb.save(reference_wb_path)
    print('Added {} new sheets to the workbook'.format(sheet_count))


def update_workbook_days_till_expiration(reference_wb_path, data_start_row, date_col, calculation_col):
    '''
    Updates each option sheet in the workbook to contain the days till expiration for that sheets option
    '''
    #loads the workbook
    wb = openpyxl.load_workbook(reference_wb_path)
    #converts date_col and calculation_col if passed as letters
    date_col=convert_to_numbers(date_col)
    calculation_col = convert_to_numbers(calculation_col)
    #loop through all the sheets in the workbook
    for (index, sheet_name) in enumerate(wb.get_sheet_names()):
        #if the sheet_name matches an option sheet:
        if re.match(OPTION_SHEET_PATTERN_INT, sheet_name) or re.match(OPTION_SHEET_PATTERN_FLOAT, sheet_name):
            #get the sheet
            sheet = wb.get_sheet_by_name(sheet_name)
            update_sheet_days_till_expiration(reference_sheet= sheet, data_start_row= data_start_row, 
                                                date_col= date_col, calculation_col= calculation_col)
    #save changes
    wb.save(reference_wb_path)
 

def update_sheet_days_till_expiration(reference_sheet, data_start_row, date_col, calculation_col):
    '''
    loops over each row containing option data and returns the days till expiration in the designated column
    '''
    #sets the total rows of the worksheet
    total_rows = reference_sheet.max_row
    #sets the expiratioin date
    exp_date = reference_sheet['B4'].value
    #sets the header of the column
    reference_sheet.cell(row=data_start_row-1, column=calculation_col).value = 'DTE'
    #loops through each row from data_start_row till total_rows
    for i in range(data_start_row, total_rows+1):
        if reference_sheet.cell(row=i, column=date_col).value == None:
            break
        else:
            curr_date = reference_sheet.cell(row=i, column=date_col).value
            reference_sheet.cell(row=i, column=calculation_col).value = days_till_expiration(start_date=curr_date, 
                                                                                            expiration_date=exp_date)

#########################Tested
def days_till_expiration(start_date, expiration_date):
    '''
    Given an expiration date and a a starting date, the days to expiration is calculated
    '''
    return (expiration_date-start_date).days

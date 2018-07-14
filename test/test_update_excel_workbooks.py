import os
import sys
import unittest
import random
import datetime as dt
from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.abspath(os.pardir)
path = os.path.join(BASE_DIR,'ma_option_vol')
sys.path.append(path)

from data_workbooks import Data_WorkSheet, Merger_Sample_Data
import update_excel_workbooks as uxlw

import base_test
def setUpModule():
	base_test.setUpModule()

def tearDownModule():
	base_test.tearDownModule()

class Test_Update_Excel_Workbooks(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.test_dir_path = base_test.TEST_DIR_PATH
		cls.acquirer_path = base_test.TEST_ACQUIRER_PATH
		cls.target_path = base_test.TEST_TARGET_PATH

	def setUp(self):
		pass

	def tearDown(slef):
		base_test.clean_up()

	def test_format_option_description_call(self):
		security_name = 'BBG0079FD3F4 Equity'	
		option_description = 'PFE US 11/28/14 C21'
		expiration_date = dt.datetime.strptime('11/28/14','%m/%d/%y').date()
		expected = option_data_list = [security_name, option_description, 'Call', expiration_date, 21]
		formated = uxlw.format_option_description(security_name, option_description)
		self.assertEqual(expected, formated)

	def test_format_option_description_put(self):
		security_name = 'BBG0079FD471 Equity'	
		option_description = 'PFE US 11/28/14 P22'
		expiration_date = dt.datetime.strptime('11/28/14','%m/%d/%y').date()
		expected = option_data_list = [security_name, option_description, 'Put', expiration_date, 22]
		formated = uxlw.format_option_description(security_name, option_description)
		self.assertEqual(expected, formated)


	def test_copy_data_single_value(self):
		wb1 = Workbook()
		wb2 = Workbook()
		sheet1 = wb1.active
		sheet2 = wb2.active
		sheet1.append(['test_value'])
		uxlw.copy_data(sheet1, sheet2,1, 1, 1, 1)
		self.assertEqual(sheet1['A1'].value , 'test_value')
		self.assertEqual(sheet2['A1'].value , 'test_value')

	def test_copy_data_column(self):
		wb1 = Workbook()
		wb2 = Workbook()
		sheet1 = wb1.active
		sheet2 = wb2.active
		for item in [['a'], ['b'], ['c'], ['d']]:
			sheet1.append(item)
		uxlw.copy_data(sheet1, sheet2,1, 4, 1, 2)
		for i, item in enumerate([['a'], ['b'], ['c'], ['d']], start=1):
			self.assertEqual(sheet2.cell(row=i, column=2).value, item[0])

	def test_days_till_expiration(self):
		start = dt.date.today()
		end = start + dt.timedelta(days=2)
		dte = uxlw.days_till_expiration(start, end)
		self.assertEqual(dte, 2)


	def test_delete_option_sheets(self):
		wb = Workbook()
		#assert that it only has one sheet
		self.assertEqual(len(wb.sheetnames), 1)
		sheet_titles = ['NEE US 12-21-13 P55', 'NEE US 12-21-13 C55', 'NEE US 12-21-13 P60', 'NEE US 12-21-13 C60']
		for item in sheet_titles:
			ws = wb.create_sheet()
			ws.title = item
		#assert that new sheets were added
		self.assertGreater(len(wb.sheetnames), 1)
		path = os.path.join(self.target_path, 'test_delete.xlsx')
		wb.save(path)
		
		#call the delete function
		uxlw.delet_workbook_option_sheets(path)
		wb = load_workbook(path)
		self.assertEqual(len(wb.sheetnames), 1)


	def test_find_index_0(self):
		#get workbook
		wb = Workbook()
		#get worksheet
		sheet = wb.active

		#Some arbitrary start date
		start_date = dt.datetime(year=2016, month=3, day=1)
		#append dates starting from cell A1
		for i in range(360):
			sheet.append([start_date + dt.timedelta(days=i)])

		#the specific date we're looking for, which is 100 spots from the starting point
		specific_date = start_date + dt.timedelta(days=100)

		index = uxlw.find_index_0(sheet,0, sheet.max_row, 1, specific_date)
		self.assertEqual(index, 101)


	def test_fill_sheet_empty_cells(self):
		wb = Workbook()
		sheet = wb.active
		data = [[1, 2,3,4,5],[2],[3],[4],[5],]
		for item in data:
			sheet.append(item)

		uxlw.fill_option_sheet_empty_cells(sheet, 1, 1, 'test')
		#check to see that empty cells have been filled in properly
		for i in range(2,sheet.max_column+1):
			for j in range(2,sheet.max_row+1):
				self.assertEqual(sheet.cell(row=j, column=i).value, 'test')

	def test_fill_wb_empty_cells(self):
		wb = Workbook()
		sheet_titles = ['NEE US 12-21-13 P55', 'NEE US 12-21-13 C55', 'NEE US 12-21-13 P60', 'NEE US 12-21-13 C60']
		for i in range(4):
			wb.create_sheet(title=sheet_titles[i])
		#remove the standard sheet
		del wb['Sheet']	
		data = [[1, 2,3,4,5],[2],[3],[4],[5],]
		#for every sheet append the data
		for sheet in wb.sheetnames:
			ws = wb[sheet]
			for item in data:
				ws.append(item)
		path = os.path.join(self.target_path, 'test_fill_wb.xlsx')
		wb.save(path)
		#function args: reference_wb_path, column_start, row_start, fill_value
		uxlw.fill_option_wb_empty_cells(path, 1, 1, 'test')
		wb = load_workbook(path)
		for sheet in wb.sheetnames:
			ws = wb[sheet]
		for i in range(2,ws.max_column+1):
			for j in range(2,ws.max_row+1):
				self.assertEqual(ws.cell(row=j, column=i).value, 'test')

	def test_is_in_range(self):
		self.assertTrue(uxlw.is_in_range(4, 5, 3))
		self.assertFalse(uxlw.is_in_range(4, 10, 6))


	def test_update_sheet_with_BDP(self):
		#workbook path
		path = os.path.join(
				os.path.dirname(os.path.abspath(__file__)),
				'samples',
				'Option_BDP_Description_Sample.xlsx')
		wb1 = load_workbook(path)
		ws1 = wb1['Options Chain']

		#making a copy of the sample worksheet to manipulate
		wb2 = Workbook()
		ws2 = wb2.active
		ws2.title = 'Options Chain'
		
		# uxlw.copy_data(reference_sheet, main_sheet,index_start_row, index_end_row, reference_data_column, main_data_column)
		for i in range(ws1.max_column):
			uxlw.copy_data(ws1, ws2, 1, ws1.max_row, i+1, i+1)

		new_path = os.path.join(self.target_path, 'Test_BDP_Description.xlsx')
		wb2.save(new_path)
		
		#update_sheet with the BDP function
		uxlw.update_sheet_with_BDP_description(new_path,'Options Chain', 1, 10)

		#load the workbook again
		wb = load_workbook(new_path)
		ws = wb['Options Chain']
		unique_tickers = []
		for i in range(1, ws.max_column+1, 2):
			for j in range(10, ws.max_row+1):
				cell = ws.cell(row=j, column=i).value
				if cell not in unique_tickers and cell != None:
					adj_cell = (ws.cell(row=j, column=i+1)).value
					self.assertIn('=BDP(', adj_cell)
					unique_tickers.append(cell)
				elif cell == None:
					break


	def test_update_options_contract_sheets(self):
		#load the sample workbook
		path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples',
			'test_stock_and_option_sheet.xlsx',
			)
		wb = load_workbook(path)
		self.assertTrue(len(wb.sheetnames)==2)
		new_path = os.path.join(self.target_path, 'add_options.xlsx')
		wb.save(new_path)

		data_table_index=['INDEX','DATE']
		data_table_header=['PX_LAST','PX_BID','PX_ASK','PX_VOLUME','OPEN_INT', 'IVOL']
		#test the update_options_contract_sheets function.
		#adds a new sheet for each option contract listed in the Options Chain sheet and pulls bloomberg data for each field listed in 
		uxlw.update_option_contract_sheets(workbook_path=new_path, 
											sheet_name='Options Chain',
											starting_col=1,
											starting_row=10,
											sheet_start_date_cell='B7',
											sheet_announce_date_cell='B8',
											sheet_end_date_cell='B9',
											data_header_row=8,
											data_table_index=data_table_index,
											data_table_header=data_table_header,
											BDH_optional_arg=['Days', 'Fill'],
											BDH_optional_val=['T','0'])

		wb = load_workbook(new_path)
		self.assertTrue(len(wb.sheetnames) > 2)

		#randomly select an option sheet, exclude stock and option chain sheet
		random_ws = wb[random.choice(wb.sheetnames[2:])]
		#test if the datat index and header were assigned properly
		header = data_table_index + data_table_header
		for i in range(random_ws.max_column):
			cell = random_ws.cell(row=8, column=i+1).value
			self.assertEqual(cell, header[i])

		cell = random_ws['B9'].value
		self.assertIn('=BDH(', cell)


	def test_update_sheet_index(self):
		path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples',
			'test_stock_and_option_sheet.xlsx',
			)
		wb = load_workbook(path)
		new_path = os.path.join(self.target_path, 'add_stock_index.xlsx')
		wb.save(new_path)
		#load the copy of the sample wb
		wb = load_workbook(new_path)
		ws = wb[wb.sheetnames[1]]
		date = ws['B4'].value
		# update_sheet_index(reference_sheet, date, start_row)
		start_row = 9
		uxlw.update_sheet_index(ws, date, start_row)
		# find_index_0(worksheet,start, end, date_col, date_0)
		index_0 = uxlw.find_index_0(ws, start_row, ws.max_row, 2, date)
		#check if the index was added correctly
		for i in range(start_row, ws.max_row+1):
			#index column is column 1
			cell = ws.cell(row=i, column=1).value
			self.assertEqual(cell, i-index_0)
		


	def test_update_workbook_index(self):
		path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples',
			'test_stock_and_option_sheet.xlsx',
			)
		wb = load_workbook(path)
		new_path = os.path.join(self.target_path, 'add_wb_index.xlsx')
		wb.save(new_path)
		#add new ws to the wb
		data_table_index=['INDEX','DATE']
		data_table_header=['PX_LAST','PX_BID','PX_ASK','PX_VOLUME','OPEN_INT', 'IVOL']
		uxlw.update_option_contract_sheets(workbook_path=new_path, 
											sheet_name='Options Chain',
											starting_col=1,
											starting_row=10,
											sheet_start_date_cell='B7',
											sheet_announce_date_cell='B8',
											sheet_end_date_cell='B9',
											data_header_row=8,
											data_table_index=data_table_index,
											data_table_header=data_table_header,
											BDH_optional_arg=['Days', 'Fill'],
											BDH_optional_val=['T','0'])
		
		#update the index for each sheet in relation to the announcement date
		uxlw.update_workbook_data_index(workbook_path=new_path, data_start_row=9, index_column='A')
		#load_workbook
		wb = load_workbook(new_path)
		ws = wb['Options Chain']
		stock_sheet = wb[wb.sheetnames[1]]
		#get the announcement date from the 'Options Chain'
		announce_date = ws['B5'].value
		index = uxlw.find_index_0(stock_sheet,9, stock_sheet.max_row, date_col=2, date_0=announce_date)
		#pick a random ws from [1:]
		random_sheet = wb[random.choice(wb.sheetnames[2:])]
		#check if the index was added correctly
		for i in range(9, random_sheet.max_row+1):
			#index column is column 1
			cell = random_sheet.cell(row=i, column=1).value
			self.assertEqual(cell, i-index)


	def test_find_column_index_by_header(self):
		path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples',
			'test_stock_and_option_sheet.xlsx',
			)
		#test the find_column_index_by_header() function
		wb = load_workbook(path)
		#test individual column
		data_dict= uxlw.find_column_index_by_header(reference_wb=wb, column_header='PX_LAST', header_row=8)
		for index, key in enumerate(data_dict):
			self.assertEqual(data_dict[key], [3])
		#test multiple columns
		data_dict= uxlw.find_column_index_by_header(reference_wb=wb, column_header=['PX_LAST', 'DATE'], header_row=8)
		for index, key in enumerate(data_dict):
			self.assertEqual(data_dict[key], [3, 2])
		

	@unittest.skip('Not Testing')
	def test_update_workbook_avg_col(self):
		#test the update_workbook_average_column() function
		#uxlw.update_workbook_average_column(reference_wb_path = test_path3, column_header='PX_LAST', header_row=8, data_start_row=9, ignore_sheet_list=['Stock Price'])
		pass


	def test_update_stock_price_sheet(self):
		path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples',
			'test_stock_and_option_sheet.xlsx',
			)
		wb = load_workbook(path)
		#delete the stock sheet thats already there
		del wb[wb.sheetnames[1]]
		self.assertEqual(len(wb.sheetnames), 1)
		new_path = os.path.join(self.target_path, 'add_wb_index.xlsx')
		#save to the updted path
		wb.save(new_path)
		#test the update_stock_price_sheet()
		data_table_index=['INDEX','DATE']
		data_table_header=['PX_LAST'] 
		uxlw.update_stock_price_sheet(workbook_path =new_path,
									sheet_name='Options Chain',
									stock_sheet_index = 1,
									sheet_start_date_cell='B7',
                                    sheet_announce_date_cell='B8', 
                                    sheet_end_date_cell='B9',  
									data_header_row=8, 
									data_table_index=data_table_index, 
									data_table_header=data_table_header, 
									BDH_optional_arg=None, 
									BDH_optional_val=None )
		#load the wb after adding stock sheet
		wb = load_workbook(new_path)
		self.assertEqual(len(wb.sheetnames), 2)
		#get the stock sheet
		stock_sheet = wb[wb.sheetnames[1]]
		meta = ['Company Name', 'Company Ticker', 'Start Date', 
				'Announcement Date', 'End Date']
		for i, item in enumerate(meta, start=1):
			cell = stock_sheet.cell(row=i, column=1).value
			self.assertEqual(cell, meta[i-1])
		cell = stock_sheet['B9'].value
		self.assertIn('=BDH(', cell)
	

	def test_data_average(self):
		lst1 = [1,2,3,4,5]
		result = uxlw.data_average(lst1)
		self.assertEqual(result, 3)

		lst2 = [1.2, 2.5, 6.7, 34.6, 22.58]
		result = uxlw.data_average(lst2)
		self.assertEqual(result, 13)


	def test_data_standard_dev(self):
		lst1 = [1,2,3,4,5]
		result = uxlw.data_standard_dev(lst1)
		self.assertEqual(result, 2)

		lst2 = [1.2, 2.5, 6.7, 34.6, 22.58]
		result = uxlw.data_standard_dev(lst2)
		self.assertEqual(result, 15)


	def stock_data_to_list(self):
		path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples',
			'test_stock_and_option_sheet.xlsx',
			)
		wb = load_workbook(path)
		ws = wb[wb.sheetnames[1]]

		stock_price_list = []
		for i in range(9, ws.max_row+1):
			cell = ws.cell(row=i, column=3).value
			stock_price_list.append(cell)

		stock_data = uxlw.stock_data_to_list(reference_wb_path=path, price_column_header='PX_LAST', header_start_row=8)
		self.assertEqual(stock_data, stock_price_list)


	def test_historic_stock_mean_and_std(self):
		path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples',
			'test_stock_and_option_sheet.xlsx',
			)
		date = dt.datetime.strptime('20140708', '%Y%m%d')
		hm_std=uxlw.historic_stock_mean_and_std(path, 'PX_LAST', header_start_row=8, date_0=date)
		self.assertEqual(hm_std, (49, 4))
		

	def test_merger_stock_mean_and_std(self):
		path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples',
			'test_stock_and_option_sheet.xlsx',
			)
		date = dt.datetime.strptime('20140708', '%Y%m%d')
		mm_std=uxlw.merger_stock_mean_and_std(path, 'PX_LAST', header_start_row=8, date_0=date)
		self.assertEqual(mm_std, (55, 3))


	def test_convert_to_numbers_single(self):
		lst = 'A'
		value = uxlw.convert_to_numbers(lst)
		self.assertEqual(value, 1)
		lst = 'B'
		value = uxlw.convert_to_numbers(lst)
		self.assertEqual(value, 2)


	def test_convert_to_numbers_list(self):
		lst = ['A', 'B', 'D', 'G', 'H', 'E', 'AA']
		value = uxlw.convert_to_numbers(lst)
		expected = [1, 2, 4, 7, 8, 5, 27]
		self.assertEqual(value, expected)

	
	@unittest.skip('Not Testing')
	def test_add_extra_sheets(self):
		# uxlw.add_extra_sheets(reference_wb_path=Pfizer_test_path, sheet_name='Options Chain', ticker_column=1, 
		# 	description_column=2,sheet_start_date_cell='B7', sheet_announce_date_cell='B8', 
		# 	sheet_end_date_cell='B9',  data_header_row=8, data_table_index=['INDEX','DATE'], 
		# 	data_table_header=['PX_LAST'], BDH_optional_arg=['Days', 'Fill'], BDH_optional_val=['W','0'])
		pass




if __name__ == '__main__':
	unittest.main()


















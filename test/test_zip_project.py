import os
import sys
import unittest
import zipfile
from unittest.mock import MagicMock
import random
import datetime as dt
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = os.path.abspath(os.pardir)
path = os.path.join(BASE_DIR,'ma_option_vol')
sys.path.append(path)

import zip_project as zp

import base_test

def setUpModule():
	base_test.setUpModule()

def tearDownModule():
	# import pdb; pdb.set_trace()
	base_test.tearDownModule()

def touch_txt(path):
	''' creates blank text files'''
	if path[-4:]!='.txt': raise ValueError('Must end in .txt')
	with open(path, 'a'):
		pass

class Test_Zip_File(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.test_file_path = base_test.TEST_TARGET_PATH
		cls.zip_file_path = base_test.TEST_ACQUIRER_PATH
		wb = Workbook()
		ws = wb.active
		ws['A1'].value = 'Test'
		cls.test_files = ['test1.xlsx', 'test2.xlsx', 'test3.xlsx', 'no1.txt', 'no2.txt']
		for file in cls.test_files:
			if file[-5:] == '.xlsx':
				wb.save(os.path.join(cls.test_file_path, file))
			else:
				touch_txt(os.path.join(cls.test_file_path, file))

	@classmethod
	def tearDownClass(cls):
		# import pdb; pdb.set_trace()
		pass

	def test_mock_files(self):
		self.assertEqual(len(os.listdir(self.test_file_path)), 5)
		for file in self.test_files:
			self.assertIn(file, os.listdir(self.test_file_path))
		self.assertEqual(len(zp.get_dir_xlsx_files(self.test_file_path)), 3)

	def test_zip_xlsx_files(self):
		file_name = 'test.zip'
		#creates the zipfile
		zip_file = zp.zip_xlsx_files(self.test_file_path, 
					os.path.join(self.zip_file_path, file_name))
		#checks that all the .xslx files are present in the zipfile
		self.assertEqual(len(zip_file.namelist()), len(zp.get_dir_xlsx_files(self.test_file_path)))
		#checks that all the .xslx files are present in the zipfile
		for file in zp.get_dir_xlsx_files(self.test_file_path):
			f_name = file.split('/')[-1]
			self.assertIn(f_name, zip_file.namelist())

		#checks that the new zip file was added to the correct directory
		self.assertIn(file_name, os.listdir(os.path.dirname(zip_file.filename)))


	def test_zip_xlsx_files_fail(self):
		err_msg = 'zip_path must include the full path and file name to store the zip file'
		with self.assertRaises(ValueError)  as err:
			file_name = 'test'
			zp.zip_xlsx_files(self.test_file_path, 
			os.path.join(self.zip_file_path, file_name))
		self.assertEqual(str(err.exception), err_msg)

	def test_zip_preserve_value(self):
		pass


if __name__ == '__main__':
	unittest.main()

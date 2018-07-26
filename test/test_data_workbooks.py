import os
import sys
import unittest
from unittest.mock import MagicMock
import random
import datetime as dt
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = os.path.abspath(os.pardir)
path = os.path.join(BASE_DIR,'ma_option_vol')
sys.path.append(path)

#class imports
from data_workbooks import (Data_WorkSheet, Merger_Sample_Data, Treasury_Sample_Data, 
							VIX_Sample_Data, Option_Workbook, Option_Chain_Sheet, 
							Stock_Sheet,Option_Sheet)	
#decorator imports
from data_workbooks import (is_row_idx_in_range, is_col_idx_in_range, 
							has_stock_sheet, has_option_sheets)

from options import Option

import base_test

def setUpModule():
	base_test.setUpModule()

def tearDownModule():
	base_test.tearDownModule()


@is_row_idx_in_range
def row_index_func(data_ws, row):
	''' mock function to test the decorators'''
	return 10

@is_col_idx_in_range
def col_index_func(data_ws, col):
	''' mock function to test the decorators'''
	return 10

# @unittest.skip('Tested')
class Test_Data_Worksheet_Decorator_Behavior(unittest.TestCase):
	@classmethod
	def setUpClass(self):
		self.mock_data_ws = MagicMock(spec=Data_WorkSheet)
		#set the row values
		self.mock_data_ws.ws_length = 4
		self.in_row_range = 3
		self.out_row_range = 5
		#set the column values
		self.mock_data_ws.ws_width = 5
		self.in_col_range = 4
		self.out_col_range = 6
		#error messages
		self.row_err_msg = f'''The sheet has {self.mock_data_ws.ws_length} rows. Unable to get data from row {self.out_row_range}'''
		self.col_err_msg =f'''The sheet has {self.mock_data_ws.ws_width} columns. Unable to get data from column {self.out_col_range}'''

	def test_is_row_idx_in_range_pass(self):
		self.assertTrue(self.in_row_range <= self.mock_data_ws.ws_length)
		val = row_index_func(self.mock_data_ws, row=self.in_row_range)
		self.assertEqual(val, 10)

	def test_is_row_idx_in_range_fail(self):
		self.assertTrue(self.out_row_range >= self.mock_data_ws.ws_length)
		with self.assertRaises(IndexError) as err:
			row_index_func(self.mock_data_ws, row=self.out_row_range)
		self.assertEqual(str(err.exception), self.row_err_msg)

	def test_is_col_idx_in_range_pass(self):
		self.assertTrue(self.in_col_range <= self.mock_data_ws.ws_width)
		val = col_index_func(self.mock_data_ws, col=self.in_col_range)
		self.assertEqual(val, 10)

	def test_is_col_idx_in_range_fail(self):
		self.assertTrue(self.out_col_range >= self.mock_data_ws.ws_width)
		with self.assertRaises(IndexError) as err:
			col_index_func(self.mock_data_ws, col=self.out_col_range)
		self.assertEqual(str(err.exception), self.col_err_msg)

# @unittest.skip('Tested')
class Test_Data_WorkSheet(unittest.TestCase):
	def setUp(self):
		self.path = 'samples/test_sample.xlsx'
		self.sheet_name = 'Filtered Sample Set'
		self.wb = load_workbook(self.path)
		self.ws = self.wb[self.sheet_name]
		self.data_ws = Data_WorkSheet(self.wb, self.sheet_name)

	def tearDown(slef):
		base_test.clean_up()

	def test__str__(self):
		self.assertEqual(self.data_ws.__str__(), self.sheet_name)

	def test_ws_length_property(self):
		self.assertEqual(self.ws.max_row, self.data_ws.ws_length)

	def test_ws_width_property(self):
		self.assertEqual(self.ws.max_column, self.data_ws.ws_width)		

	def test_headers_property(self):
		headers = []
		for i in range(self.ws.max_column):
			data = self.ws.cell(row=1, column=i+1).value
			headers.append(data)
		self.assertEqual(headers, self.data_ws.headers)
		self.assertEqual(headers, self.data_ws.headers)

	def test_row_values(self):
		row = 2
		self.assertTrue(1 <= row <= self.data_ws.ws_length)
		result = self.data_ws.row_values(row=row)
		self.assertIsInstance(result,dict)
		for key, value in result.items():
			self.assertIsInstance(key, int)
			self.assertEqual(list(value.keys()), ['column', 'value'])

	def test_row_values_exception(self):
		max_idx = self.data_ws.ws_length
		invalid_index = max_idx+1
		expected_msg = f'The sheet has {max_idx} rows. Unable to get data from row {invalid_index}'
		with self.assertRaises(IndexError) as err:
			self.data_ws.row_values(row=invalid_index)
		self.assertEqual(str(err.exception), expected_msg)

	def test_get_value(self):
		for i in range(20):
			row = random.choice(range(1, self.data_ws.ws_length+1))
			col = random.choice(range(1, self.data_ws.ws_width+1))
			value = self.data_ws.get_value(row=row, col=col)
			cell_value = self.data_ws.ws.cell(row=row, column=col).value
			self.assertEqual(cell_value, value)

	def test_get_value_row_exception(self):
		idx = self.data_ws.ws_length+1
		f'''The sheet has {self.data_ws.ws_length} rows. Unable to get data from row {idx}'''
		with self.assertRaises(IndexError):
			self.data_ws.get_value(row=idx, col=2)

	def test_get_value_col_exception(self):
		idx = self.data_ws.ws_width+1
		f'''The sheet has {self.data_ws.ws_width} columns. Unable to get data from column {idx}'''
		with self.assertRaises(IndexError):
			self.data_ws.get_value(row=2, col=idx)

	def test_get_coordiante(self):
		self.assertEqual(self.data_ws.get_coordinate(1,1), 'A1')
		self.assertEqual(self.data_ws.get_coordinate(2,1), 'A2')
		self.assertEqual(self.data_ws.get_coordinate(7,8), 'H7')

	def test_set_value(self):
		self.assertEqual(self.data_ws.ws.cell(row=10, column=10).value, None)
		self.data_ws.set_value(10, 10, "Test Value")
		self.assertEqual(self.data_ws.get_value(row=10, col=10), "Test Value")

	def test_row_index_by_date_found(self):
		date = dt.datetime.strptime('05/02/2014', '%m/%d/%Y')
		date_col = 2
		idx = self.data_ws.row_index_by_date(date, date_col)
		self.assertEqual(idx, 2)

	def test_row_index_by_date_not_found(self):
		date = dt.datetime.strptime('07/02/2014', '%m/%d/%Y')
		date_col = 2
		with self.assertRaises(IndexError) as err:
			self.data_ws.row_index_by_date(date, date_col)
			self.assertEqual(err.msg, 'Date not found')

	def test_letter_to_col_index(self):
		result = self.data_ws.letter_to_col_index('A','B','C')
		self.assertEqual(result, [1,2,3])

		result = self.data_ws.letter_to_col_index('AZ')
		self.assertEqual(result, [52])

# @unittest.skip('Tested')
class Test_Data_Worksheet_copy_data(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.test_path = base_test.TEST_TARGET_PATH

	def setUp(self):
		wb1 = Workbook()
		ws1 = wb1.active
		ws1.title = 'test'
		path1 = os.path.join(self.test_path, 'test1.xlsx')
		# print(wb1.sheetnames)
		wb1.save(path1)
		self.test_data_ws = Data_WorkSheet(load_workbook(path1), 'test')
		self.data_ws = Data_WorkSheet(load_workbook('samples/test_sample.xlsx'), 
													'Filtered Sample Set')
	def tearDown(slef):
		base_test.clean_up()

	def test_copy_data_entire_column(self):
		self.test_data_ws.copy_data(self.data_ws, 1, 1)
		# check that values were copied over correctly
		for i in range(1, self.data_ws.ws_length+1):
			ref_val = self.data_ws.get_value(row=i, col=1)
			test_val = self.test_data_ws.get_value(row=i, col=1)
			self.assertEqual(ref_val, test_val)

	def test_copy_data_single(self):
		self.test_data_ws.copy_data(self.data_ws, 1, 1, 1, 1)
		self.assertEqual(self.test_data_ws.ws['A1'].value, self.data_ws.ws['A1'].value)

	def test_copy_select_rows(self):
		self.test_data_ws.copy_data(self.data_ws, 1, 1, 2, 4)
		for i in range(2, 5):
			ref_val = self.data_ws.get_value(row=i, col=1)
			test_val = self.test_data_ws.get_value(row=i, col=1)
			self.assertEqual(ref_val, test_val)

# @unittest.skip('Tested')
class Test_Merger_Sample_Data(unittest.TestCase):
	def setUp(self):
		self.path = 'samples/test_sample.xlsx'
		self.sheet_name = 'Filtered Sample Set'
		self.wb = load_workbook(self.path)
		self.data_ws = Merger_Sample_Data(self.wb, self.sheet_name)

	def tearDown(slef):
		base_test.clean_up()

	def test_row_values_all(self):
		row=3
		output = {}
		for i, head in enumerate(self.data_ws.headers, start=1):
			output[i] = {
				'column':head, 
				'value':self.data_ws.ws.cell(row=row, column=i).value,}
		#call to the function being tested		
		row_vals = self.data_ws.row_values(row)
		#quick check to see if the output is the same length
		self.assertEqual(len(output.keys()), len(row_vals.keys()))
		#check that the function output return the correct data
		for key in output.keys():
			self.assertEqual(output[key], row_vals[key])
		
	def test_row_values_include(self):
		output = {}
		row = 2
		for i, head in enumerate(self.data_ws.headers, start=1):
			output[i] = {
				'column':head, 
				'value':self.data_ws.ws.cell(row=row, column=i).value,}
		include = self.data_ws.headers[:5]
		output = {key:value for key, value in output.items() if value['column'] in include}
		values = self.data_ws.row_values(row, include)
		self.assertEqual(len(output), len(values))
		self.assertEqual(output, values)

	def test_row_values_include_random(self):
		output = {}
		row = 2
		for i, head in enumerate(self.data_ws.headers, start=1):
			output[i] = {
				'column':head, 
				'value':self.data_ws.ws.cell(row=row, column=i).value,}
		#randomly choose 5 heders
		include = random.choices(self.data_ws.headers, k=5)
		output = {key:value for key, value in output.items() if value['column'] in include}
		values = self.data_ws.row_values(row, include)
		#test that the output was filtered by the include list
		self.assertEqual(output, values)


# @unittest.skip('Tested')
class Test_Treasury_Sample_Data(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.treasury_sheet = Treasury_Sample_Data()
		cls.date = dt.datetime(year=2013, month=3, day=27)

	def test_row_index_by_date(self):
		value = self.treasury_sheet.row_index_by_date(self.date)
		self.assertEqual(value, 292)
	
	def test_rf_3m(self):
		value = self.treasury_sheet.rf_3m_on(date=self.date)	
		self.assertEqual(value, 0.00089)

	def test_rf_6m(self):
		value = self.treasury_sheet.rf_6m_on(date=self.date)
		self.assertEqual(value, 0.00109)

	def test_rf_12m(self):
		value = self.treasury_sheet.rf_12m_on(date=self.date)
		self.assertEqual(value, 0.00124)

	def test_is_negative_above(self):
		self.assertEqual(self.treasury_sheet.is_negative(10), 10)

	def test_is_negative_below(self):
		self.assertEqual(self.treasury_sheet.is_negative(-10), 0)


# @unittest.skip('Tested')
class Test_VIX_Sample_Data(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.vix_sheet = VIX_Sample_Data()
		cls.date = dt.datetime(year=2013, month=3, day=27)

	def test_row_index_by_date(self):
		value = self.vix_sheet.row_index_by_date(self.date)
		self.assertEqual(value, 292)
	
	def test_vix_on(self):
		value = self.vix_sheet.get_vix_on(date=self.date)
		self.assertEqual(value, 13.15)
		
	

# @unittest.skip('Tested')
class Test_Option_Chain_Sheet(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.test_path = base_test.TEST_TARGET_PATH
		cls.path1 = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples', 'test_stock_and_option_sheet.xlsx',)
		cls.path2 = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples', 'Option_BDP_Description_Sample.xlsx')

	def setUp(self):
		self.wb1 = load_workbook(self.path1)
		self.opt_chain1 = Option_Chain_Sheet(self.wb1)
		self.wb2 = load_workbook(self.path2)
		self.opt_chain2 = Option_Chain_Sheet(self.wb2)
		
	
	def test_option_BDP_description(self):
		#10 and 331 are row values specified in the example wb
		random_row = random.choice(range(10, 331))
		self.assertTrue(self.opt_chain2.get_value(row=random_row, col=2)==None)
		self.opt_chain2.option_BDP_description(random_row, 1)
		value = self.opt_chain2.get_value(row=random_row, col=2)
		self.assertIn('=BDP(', value)

	def test_sheet_BDP_description(self):
		#loop through the tickers and assert that they don't have a description yet
		for i in range(1, self.opt_chain2.ws_width+1, 2):
			for j in range(10, self.opt_chain2.ws_length+1):
				if i >= self.opt_chain2.ws_width:
					break
				else:
					value = self.opt_chain2.get_value(row=j, col=i+1)
					self.assertEqual(value, None)
		#run the function
		self.opt_chain2.sheet_BDP_description()
		unique = []
		#loop through again and assert that unique tickers have a description
		for i in range(1, self.opt_chain2.ws_width+1, 2):
			for j in range(10, self.opt_chain2.ws_length+1):
				ticker = self.opt_chain2.get_value(row=j, col=i) 
				value = self.opt_chain2.get_value(row=j, col=i+1)
				if ticker not in unique and ticker != None:
					self.assertIn('=BDP(', value)
					unique.append(ticker)
				else:
					self.assertEqual(value, None)


# @unittest.skip('Tested')
class Test_Option_Chain_Sheet_option_exp_in_range(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		'''
		In the test worksheet here are the values from start-end
		Start Date: 2013-07-13 Announcement Date: 2014-07-08 End Date: 2014-10-20
		'''
		cls.opt_chain = Option_Chain_Sheet(load_workbook('samples/test_stock_and_option_sheet.xlsx'))
		cls.date_in_range = dt.datetime(2013, 7, 22)
		cls.date_below_range = dt.datetime(2013, 5, 13)
		cls.date_above_range = dt.datetime(2015, 7, 13)

	def test_is_option_exp_in_range_inside(self):
		value = self.opt_chain.is_option_exp_in_range(self.date_in_range, 
											from_start=8, past_announcemt=60)
		self.assertTrue(value)

	def test_is_option_exp_in_range_below(self):
		value = self.opt_chain.is_option_exp_in_range(self.date_below_range, 
											from_start=8, past_announcemt=60)
		self.assertFalse(value)

	def test_is_option_exp_in_range_above(self):
		value = self.opt_chain.is_option_exp_in_range(self.date_above_range,
											from_start=8, past_announcemt=60)
		self.assertFalse(value)

# @unittest.skip('Tested')
class Test_Stock_Sheet(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples','test_stock_and_option_sheet.xlsx',)
		cls.wb = load_workbook(cls.path)
		cls.ws_name = cls.wb.sheetnames[1]
		cls.date = dt.datetime(year=2013,month=7, day=24)

	def setUp(self):
		self.stock_sheet = Stock_Sheet(self.wb, self.ws_name)

	def test_stock_sheet_details(self):
		self.assertEqual(self.stock_sheet.get_value(row=1,col=1), 'Company Name')
		self.assertEqual(self.stock_sheet.get_value(row=2,col=1), 'Company Ticker')
		self.assertEqual(self.stock_sheet.get_value(row=3,col=1), 'Start Date')
		self.assertEqual(self.stock_sheet.get_value(row=4,col=1), 'Announcement Date')
		self.assertEqual(self.stock_sheet.get_value(row=5,col=1), 'End Date')

	def test_row_index_by_date(self):
		value = self.stock_sheet.row_index_by_date(date=self.date)
		self.assertEqual(value, 16)

	def test_get_price_on(self):
		value = self.stock_sheet.get_price_on(date=self.date)
		self.assertEqual(value, 43.64)

	def test_get_index_on(self):
		self.stock_sheet.add_index()
		index = self.stock_sheet.get_index_on(date=self.date)
		self.assertEqual(index, -249)
		index = self.stock_sheet.get_index_on(date=self.stock_sheet.announce_date)
		self.assertEqual(index, 0)

	def test_px_last_lst(self):
		price_list = self.stock_sheet.px_last_lst()
		self.assertIsInstance(price_list, list)
		lst_len = len(price_list)
		expected_list = [self.stock_sheet.get_value(row=x, col=3) 
				for x in range(self.stock_sheet.header_index+1, self.stock_sheet.ws_length+1)
				if self.stock_sheet.get_value(row=x, col=3) != 0]
		self.assertEqual(price_list, expected_list)

	def test_merger_mean(self):
		mean = self.stock_sheet.merger_mean()
		#mean verified in excel
		self.assertEqual(mean, 55)
		self.assertIsInstance(mean, int)
		
	def test_merger_std(self):
		std = self.stock_sheet.merger_std()
		#standard deviation verified in excel
		self.assertEqual(std, 3)
		self.assertIsInstance(std, int)

	def test_historic_mean(self):
		mean = self.stock_sheet.historic_mean()
		#mean verified in excel
		self.assertEqual(mean, 49)
		self.assertIsInstance(mean, int)

	def test_historic_std(self):
		std = self.stock_sheet.historic_std()
		#standard deviation verified in excel
		self.assertEqual(std, 4)
		self.assertIsInstance(std, int)

	def test_is_strike_in_range(self):
		# def is_strike_in_range(self, stike, std_multiple=1.5):
		#we know mm=54 and ms=7, hm=47 and hs=10
		in_range, below_range, above_range = 54, 20, 80
		mm, ms = self.stock_sheet.merger_mean(), self.stock_sheet.merger_std()
		hm, hs = self.stock_sheet.historic_mean(), self.stock_sheet.historic_std()
		self.assertTrue(self.stock_sheet.is_strike_in_range(in_range, mm, ms, hm, hs))
		self.assertFalse(self.stock_sheet.is_strike_in_range(below_range, mm, ms, hm, hs))
		self.assertFalse(self.stock_sheet.is_strike_in_range(above_range, mm, ms, hm, hs))

	def test_add_index(self):
		#show that before the function is run all values are None
		for i in range(10, self.stock_sheet.ws_length+1):
			self.assertEqual(self.stock_sheet.get_value(row=i, col=1), None)
		self.stock_sheet.add_index()
		index = self.stock_sheet.row_index_by_date(self.stock_sheet.announce_date)
		#show that the index has been added after running the function
		for i in range(10, self.stock_sheet.ws_length+1):
			value = self.stock_sheet.get_value(row=i, col=1)
			self.assertIsInstance(value, int)
			self.assertEqual(value, i-index)

# @unittest.skip('Tested')
class Test_Option_Sheet(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.test_path = os.path.join(base_test.TEST_TARGET_PATH, 'test_option_sheet.xlsx')
		cls.path1 = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples','test_stock_and_option_sheet.xlsx',)
		cls.wb = load_workbook(cls.path1)
		cls.wb.save(cls.test_path)
		cls.option_wb = Option_Workbook(cls.test_path)
		cls.stock_sheet = cls.option_wb.stock_sheet
		cls.stock_sheet.add_index()
		cls.chain_sheet = Option_Chain_Sheet(cls.option_wb.wb)
		ticker = cls.chain_sheet.get_value(row=10,col=1)
		description = cls.chain_sheet.get_value(row=10,col=2)
		cls.option_wb.create_option_sheet(ticker, description)
		sheet = cls.option_wb.option_sheetnames[0]
		cls.option_sheet = Option_Sheet(cls.option_wb.wb, sheet)
		start = cls.chain_sheet.start_date
		row = 9
		while start <= cls.option_sheet.ws['B4'].value:
			cls.option_sheet.set_value(row=row, col=2, value=start)
			row+=1
			start+=dt.timedelta(days=1)
		cls.option_wb.save()
		cls.treasury_sheet = Treasury_Sample_Data()
		cls.date = cls.stock_sheet.ws['B12'].value


	def test_option_sheet_details(self):
		#test the metadata headers
		self.assertEqual(self.option_sheet.get_value(row=1,col=1), 'Ticker')
		self.assertEqual(self.option_sheet.get_value(row=2,col=1), 'Description')
		self.assertEqual(self.option_sheet.get_value(row=3,col=1), 'Type')
		self.assertEqual(self.option_sheet.get_value(row=4,col=1), 'Expiration Date')
		self.assertEqual(self.option_sheet.get_value(row=5,col=1), 'Strike Price')
		
	def test_fill_empty_cells(self):
		empty_cells = {'count':0, 'cells':[]}
		for i in range(3, self.option_sheet.ws_width+1):
			for j in range(self.option_sheet.header_index+1, self.option_sheet.ws_length+1):
				value = self.option_sheet.get_value(row=j, col=i)
				if value == None:
					empty_cells['count']+=1
					empty_cells['cells'].append(self.option_sheet.get_coordinate(row=j, col=i))
		self.assertTrue(empty_cells['count'] > 0)
		fill_value = 0
		self.option_sheet.fill_empty_cells(fill_value=fill_value)
		for cell in empty_cells['cells']:
			self.assertTrue(self.option_sheet.ws[cell].value == fill_value)

	def test_get_stock_index(self):
		#self.stock_sheet
		date = self.stock_sheet.ws['B13'].value
		index = self.option_sheet.get_stock_index(self.stock_sheet, date)
		self.assertEqual(index, -252)

	def test_copy_index(self):
		#show that the index is blank
		for i in range(self.option_sheet.header_index+1, self.option_sheet.ws_length+1):
			option_index = self.option_sheet.ws.cell(row=i, column=1).value
			self.assertEqual(option_index, None)
		#run the function:
		self.option_sheet.copy_index(self.stock_sheet)
		for i in range(self.option_sheet.header_index+1, self.option_sheet.ws_length+1):
			option_index = self.option_sheet.get_value(row=i, col=1)
			date = self.option_sheet.get_value(row=i, col=self.option_sheet.date_col)
			if option_index != 'N/A':
				stock_index = self.stock_sheet.get_index_on(date=date)
				self.assertEqual(option_index, stock_index)

	def test_get_stock_price_valid_date(self):
		date = self.stock_sheet.ws['B14'].value
		price = self.option_sheet.get_stock_price(self.stock_sheet, date)
		self.assertIsInstance(price, float)
		self.assertEqual(price, 44.97)

	def test_get_stock_price_invalid_date(self):
		date = dt.datetime(year=2019, month=1, day=1)
		value = self.option_sheet.get_stock_price(self.stock_sheet, date)
		#because the date is not in the sheet, 0 should be returned
		self.assertEqual(value, 0)
				
	def test_get_risk_free_rate_3_6_or_12(self):
		for rate in [3, 6, 12]:
			rf = self.option_sheet.get_risk_free_rate(self.treasury_sheet, self.date, rate)
			self.assertIsInstance(rf, float)

	def test_get_risk_free_rate_wrong_num(self):
		message = 'Rate must be set to either 3, 6, or 12'
		with self.assertRaises(ValueError)as err:
			self.option_sheet.get_risk_free_rate(self.treasury_sheet, self.date, 10)
		self.assertEqual( message, str(err.exception))


class Test_Option_Sheet_iv_and_vega(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.test_path = os.path.join(base_test.TEST_TARGET_PATH, 'test_option_sheet.xlsx')
		cls.path = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples','option_test.xlsx',)
		wb = load_workbook(cls.path)
		cls.wb_sheets = wb.sheetnames
		wb.save(cls.test_path)
		cls.option_workbook = Option_Workbook(cls.test_path)
		cls.stock_sheet = cls.option_workbook.stock_sheet
		cls.treasury_sheet = Treasury_Sample_Data()

	def setUp(self):
		self.option_sheet = Option_Sheet(self.option_workbook.wb, self.wb_sheets[2])

	def test_add_iv_calculation(self):
		row = 12
		date = self.option_sheet.get_value(row=row, col=2)
		col = self.option_sheet.ws_width+1
		stock_price = self.stock_sheet.get_price_on(date=date)
		rf_rate =  self.treasury_sheet.rf_6m_on(date=date)
		#assert that the cell is empty:
		self.assertEqual(self.option_sheet.ws.cell(row=row, column=col).value, None)
		#run the function
		iv = self.option_sheet.add_iv_calculation(row, col, date, stock_price, rf_rate)
		self.assertIsInstance(self.option_sheet.get_value(row=row, col=col), float)
		self.assertEqual(self.option_sheet.get_value(row=row, col=col), iv)

	def test_add_vega_calculation(self):
		row = 14
		date = self.option_sheet.get_value(row=row, col=2)
		col = self.option_sheet.ws_width+1
		stock_price = self.stock_sheet.get_price_on(date=date)
		rf_rate =  self.treasury_sheet.rf_6m_on(date=date)
		#assert that the cell is empty:
		self.assertEqual(self.option_sheet.ws.cell(row=row, column=col).value, None)
		#run the function
		vega = self.option_sheet.add_vega_calculation(row, col, date, stock_price, rf_rate)
		self.assertIsInstance(self.option_sheet.get_value(row=row, col=col), float)
		self.assertEqual(self.option_sheet.get_value(row=row, col=col), vega)

	def test_sheet_iv_calculation_correct(self):
		header_row = self.option_sheet.header_index
		col = self.option_sheet.ws_width+1
		row = self.option_sheet.header_index+1
		#loop over the date column until its empty
		while self.option_sheet.get_value(row=row, col=self.option_sheet.date_col) !=None:
			value = self.option_sheet.ws.cell(row=row, column=col).value
			self.assertEqual(value, None)
			row+=1
		#call the function
		self.option_sheet.sheet_iv_calculation(col, self.stock_sheet, 
						self.treasury_sheet, rate=3, heading='test heading')
		#assert that the heading was added correctly
		self.assertEqual(self.option_sheet.get_value(row=header_row, col=col), 'test heading')
		row = self.option_sheet.header_index+1
		#assert that the same cells checked before now have values
		while self.option_sheet.get_value(row=row, col=self.option_sheet.date_col) !=None:
			value = self.option_sheet.get_value(row=row, col=col)
			self.assertTrue(value != None)
			row+=1

	def test_sheet_iv_calculation_error(self):
		message = 'Rate must be set to either 3, 6, or 12'
		col = self.option_sheet.ws_width+1
		with self.assertRaises(ValueError)as err:
			self.option_sheet.sheet_iv_calculation(col, self.stock_sheet, 
						self.treasury_sheet,rate=10, heading='3Month IV')
		self.assertEqual( message, str(err.exception))

	def test_sheet_vega_calculation(self):
		header_row = self.option_sheet.header_index
		col = self.option_sheet.ws_width+1
		row = self.option_sheet.header_index+1
		#loop over the date column until its empty
		while self.option_sheet.get_value(row=row, col=self.option_sheet.date_col) !=None:
			value = self.option_sheet.ws.cell(row=row, column=col).value
			self.assertEqual(value, None)
			row+=1
		#call the function
		self.option_sheet.sheet_vega_calculation(col, self.stock_sheet, 
						self.treasury_sheet, rate=6, heading='test heading')
		#assert that the heading was added correctly
		self.assertEqual(self.option_sheet.get_value(row=header_row, col=col), 'test heading')
		row = self.option_sheet.header_index+1
		#assert that the same cells checked before now have values
		while self.option_sheet.get_value(row=row, col=self.option_sheet.date_col) !=None:
			value = self.option_sheet.get_value(row=row, col=col)
			self.assertTrue(value != None)
			row+=1

	def test_sheet_vega_calculation_error(self):
		message = 'Rate must be set to either 3, 6, or 12'
		col = self.option_sheet.ws_width+1
		with self.assertRaises(ValueError)as err:
			self.option_sheet.sheet_vega_calculation(col, self.stock_sheet, 
						self.treasury_sheet,rate=11, heading='Vega')
		self.assertEqual( message, str(err.exception))

# @unittest.skip('Tested')
class Test_Option_Workbook(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		cls.path1 = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples','test_stock_and_option_sheet.xlsx',)
		cls.path2 = os.path.join(
			os.path.dirname(os.path.abspath(__file__)),
			'samples','Option_BDP_Description_Sample.xlsx',)
		
	def setUp(self):
		self.opt_wb1 = Option_Workbook(self.path1)
		self.cs = self.opt_wb1.chain_sheet
		self.opt_wb2 = Option_Workbook(self.path2)
		self.option_labels = ['Ticker', 'Description', 'Type', 'Expiration Date', 'Strike Price']
		self.stock_lables = ['Company Name', 'Ticker', 'Start Date', 'Announcement Date', 'End Date']

	def tearDown(slef):
		base_test.clean_up()

	def test_stock_sheet_return_sheet(self):
		self.assertIsInstance(self.opt_wb1.stock_sheet, Stock_Sheet)

	def test_stock_sheet_return_none(self):
		self.assertEqual(self.opt_wb2.stock_sheet, None)

	def test_option_sheetnames_return_lst(self):
		ticker = self.opt_wb1.chain_sheet.get_value(row=10,col=1)
		description = self.opt_wb1.chain_sheet.get_value(row=10,col=2)
		self.opt_wb1.create_option_sheet(ticker, description)
		self.assertIsInstance(self.opt_wb1.option_sheetnames, list)
		#assert that the option_sheetnames is a subset of the wb.sheetnames list
		for sheet in self.opt_wb1.option_sheetnames:
			self.assertIn(sheet, self.opt_wb1.wb.sheetnames)

	def test_option_sheetnames_return_none(self):
		self.assertEqual(self.opt_wb2.option_sheetnames, None)

	def test_save(self):
		file_name = 'save_test.xlsx'
		test_path = os.path.join(base_test.TEST_TARGET_PATH, file_name)
		self.opt_wb1.save(test_path)
		self.assertIn(file_name, os.listdir(base_test.TEST_TARGET_PATH))

	def test_stock_meta_data(self):
		meta_data = self.opt_wb1.stock_meta_data()
		self.assertIsInstance(meta_data, list)
		data = [ self.cs.company_name, f'{self.cs.ticker} {self.cs.type}', self.cs.start_date,
				self.cs.announce_date, self.cs.end_date]
		for i, item in enumerate(meta_data):
			header, value = item
			self.assertEqual(header, self.stock_lables[i])
			self.assertEqual(value, data[i])

	def test_create_stock_sheet(self):
		self.assertEqual(len(self.opt_wb2.wb.sheetnames), 1)
		self.opt_wb2.create_stock_sheet()
		self.assertEqual(len(self.opt_wb2.wb.sheetnames), 2)
		expected_stock_sheetname = f'{self.opt_wb2.chain_sheet.ticker} {self.opt_wb2.chain_sheet.type}'
		self.assertEqual(self.opt_wb2.wb.sheetnames[1], expected_stock_sheetname)
		self.assertIsInstance(self.opt_wb2.wb[expected_stock_sheetname], Worksheet)

	def test_option_meta_data(self):
		ticker = self.cs.get_value(row=10, col=1)
		description = self.cs.get_value(row=10, col=2)
		meta_data = self.opt_wb1.option_meta_data(ticker, description)
		values = [ticker, description, 'Put', dt.datetime.strptime('07/20/13', '%m/%d/%y'), 27.5 ]
		for i, item in enumerate(meta_data):
			label, value = item
			self.assertEqual(label, self.option_labels[i])
			self.assertEqual(value, values[i])

	def test_create_option_sheet(self):
		ticker = self.cs.get_value(row=10, col=1)
		description = self.cs.get_value(row=10, col=2)
		sheet_name = description.replace('/', '-')
		self.assertEqual(self.opt_wb1.option_sheetnames, None)
		self.opt_wb1.create_option_sheet(ticker, description)
		self.assertEqual(self.opt_wb1.option_sheetnames, [sheet_name])
		self.assertIsInstance(self.opt_wb1.wb[sheet_name], Worksheet)


	def test_add_option_sheets(self):
		# self.assertEqual(self.opt_wb1.option_sheetnames, None)
		self.opt_wb1.add_option_sheets()
		self.assertIsInstance(self.opt_wb1.option_sheetnames, list)

	def test_proper_desciption_format_none(self):
		self.assertFalse(self.opt_wb1.proper_desciption_format(None))

	def test_proper_description_format_wrong_str(self):
		random_str = 'ertertert'
		self.assertFalse(self.opt_wb1.proper_desciption_format(random_str))

	def test_proper_description_format_correct(self):
		desc_int = 'PFE US 12/20/14 P18'
		self.assertTrue(self.opt_wb1.proper_desciption_format(desc_int))
		desc_flot = 'PFE US 12/20/14 C18.5'
		self.assertTrue(self.opt_wb1.proper_desciption_format(desc_flot))

	@unittest.skip('Finish writing method')
	def test_calculate_wb_iv(self):
		pass

	@unittest.skip('Finish writing method')
	def test_calculate_wb_vega(self):
		pass

@has_stock_sheet
def stock_func(opt_wb):
	''' mock function to test the decorators'''
	return 10

@has_option_sheets
def option_func(opt_wb):
	''' mock function to test the decorators'''
	return 10

# @unittest.skip('Tested')
class Test_Option_Workbook_Decorator_Behavior(unittest.TestCase):
	@classmethod
	def setUpClass(self):
		#proper return values
		self.mock_opt_wb1 = MagicMock(spec=Option_Workbook)
		self.mock_opt_wb1.option_sheetnames = ['some', 'list', 'of', 'option', 'sheetnames']
		self.mock_opt_wb1.stock_sheet = MagicMock(spec=Stock_Sheet)
		#inproper return values
		self.mock_opt_wb2 = MagicMock(spec=Option_Workbook)
		self.mock_opt_wb2.option_sheetnames = None
		self.mock_opt_wb2.stock_sheet = None
		#error messages
		self.stock_err_msg = 'Please run create_stock_sheet method'
		self.option_err_msg ='Please run add_option_sheets method'

	def test_has_stock_sheet_pass(self):
		self.assertEqual(stock_func(self.mock_opt_wb1), 10)
			
	def test_has_stock_sheet_fail(self):
		with self.assertRaises(AttributeError) as err:
			stock_func(self.mock_opt_wb2)
		self.assertEqual(str(err.exception), self.stock_err_msg)

	def test_has_option_sheets_pass(self):
		self.assertEqual(option_func(self.mock_opt_wb1), 10)

	def test_has_option_sheets_fail(self):
		with self.assertRaises(AttributeError) as err:
			option_func(self.mock_opt_wb2)
		self.assertEqual(str(err.exception), self.option_err_msg)



if __name__ == "__main__":
	unittest.main()
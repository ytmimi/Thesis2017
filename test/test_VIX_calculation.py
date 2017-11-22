#reference data is based on the article that can be found at https://www.cboe.com/micro/vix/vixwhite.pdf
#testing the VIX_calculation module is based on recreating the results produced by the CBOE
import os
import sys
import datetime as dt
import openpyxl
from test_path import vix_data
parent_path = os.path.abspath(os.pardir)
path = os.path.join(parent_path,'ma_option_vol')
#adds the file path for the ma_options_vol module to the path that python will search in order to look for modules
sys.path.append(path)
import VIX_calculation as vix_c

vix_wb = openpyxl.load_workbook(vix_data)
vix_sheet = vix_wb.get_sheet_by_name('Data Set')

#total rows in the vix sheet
total_rows = vix_sheet.max_row
#starting data row
start_row = 4

near_term_strike_column=1
near_call_bid_column=2
near_call_ask_column=3
near_put_bid_column=4
near_put_ask_column=5

next_term_strike_column=6
next_call_bid_column=7
next_call_ask_column=8
next_put_bid_column=9
next_put_ask_column=10


#dictionaries to store Option_Contract_Data objects
near_option_dict={'call':[], 'put':[]}

next_option_dict={'call':[], 'put':[]}


#loop through each row of the vix_sheet and create an instance of Option_Contract_Data
for i in range(start_row, total_rows+1):
	if vix_sheet.cell(row=i, column=near_term_strike_column).value != None:
		option1 = vix_c.Option_Contract_Data(option_description='N/A', 
											exp_date=vix_sheet.cell(row=16, column=13).value,
											strike_price=vix_sheet.cell(row=i, column=near_term_strike_column).value, 
											px_last = 'N/A', 
											px_bid=vix_sheet.cell(row=i, column= near_call_bid_column).value, 
											px_ask=vix_sheet.cell(row=i, column= near_call_ask_column).value)
		near_option_dict['call'].append(option1)
		
		option2 = vix_c.Option_Contract_Data(option_description='N/A', 
											exp_date=vix_sheet.cell(row=16, column=13).value, 
											strike_price=vix_sheet.cell(row=i, column=near_term_strike_column).value, 
											px_last = 'N/A', 
											px_bid=vix_sheet.cell(row=i, column= near_put_bid_column).value, 
											px_ask=vix_sheet.cell(row=i, column= near_put_ask_column).value)
		near_option_dict['put'].append(option2)

	if vix_sheet.cell(row=i, column=next_term_strike_column).value != None:
		option3 = vix_c.Option_Contract_Data(option_description='N/A', 
											exp_date=vix_sheet.cell(row=17, column=13).value,
											strike_price=vix_sheet.cell(row=i, column=next_term_strike_column).value, 
											px_last = 'N/A', 
											px_bid=vix_sheet.cell(row=i, column= next_call_bid_column).value, 
											px_ask=vix_sheet.cell(row=i, column= next_call_ask_column).value)
		next_option_dict['call'].append(option3)

		option4 = vix_c.Option_Contract_Data(option_description='N/A', 
											exp_date=vix_sheet.cell(row=17, column=13).value, 
											strike_price=vix_sheet.cell(row=i, column=next_term_strike_column).value, 
											px_last = 'N/A', 
											px_bid=vix_sheet.cell(row=i, column= next_put_bid_column).value, 
											px_ask=vix_sheet.cell(row=i, column= next_put_ask_column).value)
		next_option_dict['put'].append(option4)

r1 = vix_sheet.cell(row=18, column=13).value
today = vix_sheet.cell(row=15, column=13).value

#create an instance of Near_Term with the above near_option_dict
near = vix_c.Near_Term(option_dict=near_option_dict, risk_free_rate=r1, current_date= today, current_time='9:46 AM', settlement_time='8:30 AM')

print('T1: \t  {}\nexpected: 0.0683486'.format((near.T)))
print('\n')
print('R1: \t  {}\nexpected: 0.000305'.format(near.R))
print('\n')
print('F_strike: {}\nexpected: 1965'.format(near.F_strike))
print('\n')
print('Forward price:  {}\nexpected: \t1962.89996'.format(near.F))
print('\n')
print('k0: \t  {}\nexpected: 1960'.format(near.k0))
print('\n')
print('forward contribution: {}\nexpected: 0.00003203'.format(near.forward_contribution()))
print('\n')
print('variance: {}\nexpected: 0.01846292'.format(near.variance))
print('\n')
print('************')
r2 = vix_sheet.cell(row=19, column=13).value
#create an instance of Next_Term with the above next_option_dict


next_ = vix_c.Next_Term(option_dict=next_option_dict, risk_free_rate=r2, current_date= today, current_time='9:46 AM', settlement_time='3:00 PM')
print('T2: \t  {}\nexpected: 0.0882686'.format((next_.T)))
print('\n')
print('R2: \t  {}\nexpected: 0.000286'.format(next_.R))
print('\n')
print('F_strike: {}\nexpected: 1960'.format(next_.F_strike))
print('\n')
print('Forward price:  {}\nexpected: \t1962.40006'.format(next_.F))
print('\n')
print('k0: \t  {}\nexpected: 1960'.format(next_.k0))
print('\n')
print('forward contribution: {}\nexpected:  0.00001699'.format(next_.forward_contribution()))
print('\n')
print('variance: {}\nexpected: 0.01882101'.format(next_.variance))
print('************')


vix_calc = vix_c.VIX_Calculation(Near_Term=near, Next_Term=next_)
print('calculated VIX: {}'.format(vix_calc.VIX))
print('expected VIX:   13.685821')







import os
import sys
import unittest
import datetime as dt
from openpyxl import Workbook, load_workbook
BASE_DIR = os.path.abspath(os.pardir)
path = os.path.join(BASE_DIR,'ma_option_vol')
sys.path.append(path)
import generate_company_workbooks as gcw
from base_test import Test_Base

# c = gcw.Create_Company_Workbooks(source_sheet_name, source_file, target_path, acquirer_path)
# c.create_company_workbooks()


class Test_Generate_Company_WB(Test_Base):
	def setUp(self):
		#Note: note that target_path and acquirer_path are defined in super()
		self.wb = Workbook()
		self.sheet_name = 'Filtered Sample Set'
		self.source_file = 'samples/test_sample.xlsx'
		self.wb_generator = gcw.Create_Company_Workbooks(self.sheet_name, 
					self.source_file, self.target_path, self.acquirer_path)
		
	def tearDown(self):
		super().tearDown()

	def test_formated_wb_path_default(self):
		# formated_wb_path(self, file_name, date_str='', file_extension='xlsx', path=''):
		file_name = 'test_file'
		date = dt.datetime.strptime('2018-06-20', '%Y-%m-%d')
		final_file_name = self.wb_generator.formated_wb_path(file_name, date)
		self.assertEqual(final_file_name, './test_file_2018-06-20.xlsx')

	def test_formated_wb_path_kwargs(self):
		# formated_wb_path(self, file_name, date_str, file_extension='xlsx', path=''):
		file_name = 'test_file'
		date = dt.datetime.strptime('2018-06-20', '%Y-%m-%d')
		path = '/Users/username/Documents/'
		final_file_name = self.wb_generator.formated_wb_path(file_name, date, path=path)
		self.assertEqual(final_file_name, '/Users/username/Documents/test_file_2018-06-20.xlsx')

	def test_save_new_workbook_target(self):
		file_name = 'target_test'
		date = dt.datetime.strptime('2018-06-20', '%Y-%m-%d')
		# def save_new_workbook(self,new_workbook,workbook_path, file_name, start_date_str, file_extension='xlsx'):
		self.wb_generator.save_new_workbook(self.wb, file_name, date, path=self.target_path)
		self.assertIn('target', os.listdir(self.test_dir_path))
		date = date.strftime('%Y-%m-%d')
		self.assertIn(f'{file_name}_{date}.xlsx', os.listdir(self.target_path))

	
	def test_save_workbook_acquirer(self):
		file_name = 'acquirer_test'
		date = dt.datetime.strptime('2018-06-20', '%Y-%m-%d')
		self.wb_generator.save_new_workbook(self.wb, file_name, date, path=self.acquirer_path)
		self.assertIn('acquirer', os.listdir(self.test_dir_path))
		date = date.strftime('%Y-%m-%d')
		self.assertIn(f'{file_name}_{date}.xlsx', os.listdir(self.acquirer_path))

	def test_adjust_to_weekday_saturday(self):
		#check the calendar, this date is a Saturday
		saturday = dt.date(year=2018, month=1, day=6)
		date = self.wb_generator.adjust_to_weekday(saturday)
		self.assertEqual(date.weekday(), 0)

	def test_adjust_to_weekday_sunday(self):
		#check the calendar, this date is a Sunday
		sunday = dt.date(year=2018, month=1, day=7)
		date = self.wb_generator.adjust_to_weekday(sunday)
		self.assertEqual(date.weekday(), 0)
	
	def test_create_company_workbooks_target(self):	
		wb = load_workbook(self.source_file)
		sheet = wb['Filtered Sample Set']
		for (i, row) in enumerate(sheet.rows):
			if i != 0:
				target_name = row[3].value
				announcement_date = self.wb_generator.adjust_to_weekday(row[1].value.date())
				new_file = f'{target_name}_{announcement_date}.xlsx'
				self.wb_generator.new_company_workbook(i+1, self.target_path)
				self.assertIn(new_file, os.listdir(self.target_path))
		
	def test_create_company_workbooks_acquirer(self):	
		wb = load_workbook(self.source_file)
		sheet = wb['Filtered Sample Set']
		for (i, row) in enumerate(sheet.rows):
			if i != 0:
				acquirer_name = row[6].value
				announcement_date = self.wb_generator.adjust_to_weekday(row[1].value.date())
				new_file = f'{acquirer_name}_{announcement_date}.xlsx'
				self.wb_generator.new_company_workbook(i+1, self.acquirer_path)
				self.assertIn(new_file, os.listdir(self.acquirer_path))

	def test_company_workbook_data_target(self):
		#create the workbook
		wb = load_workbook(self.source_file)
		sheet = wb['Filtered Sample Set']
		rows = sheet.rows
		for i in range(2):
			rows.__next__()
		row = rows.__next__()
		start_date = (row[1].value - dt.timedelta(days=360))
		data = [
			('Target Name',row[3].value),
			('Target Ticker',row[4].value),
			('Type','Equity'),
			('Start Date', start_date),
			('Announcement Date', row[1].value),
			('End Date', row[2].value),
			('Formated Start Date', int(start_date.strftime('%Y%m%d'))),
			('Formated Announcement Date', int(row[1].value.date().strftime('%Y%m%d'))),
			('Formated End Date', int(row[2].value.strftime('%Y%m%d'))),
		]

		#create the file from source wb 
		self.wb_generator.new_company_workbook(3, self.target_path)
		#load the newly created wb
		test_wb_file = os.listdir(self.target_path)[0]
		path = os.path.join(self.target_path, test_wb_file)
		test_wb = load_workbook(path)
		ws = test_wb['Options Chain']
		for row in range(9):
			for col in range(2):
				self.assertEqual(ws.cell(column=col+1, row=row+1).value,data[row][col])

	def test_company_workbook_data_acquirer(self):
		#create the workbook
		wb = load_workbook(self.source_file)
		sheet = wb['Filtered Sample Set']
		rows = sheet.rows
		for i in range(2):
			rows.__next__()
		row = rows.__next__()
		start_date = (row[1].value - dt.timedelta(days=360))
		data = [
			('Acquirer Name',row[6].value),
			('Acquirer Ticker',row[7].value),
			('Type','Equity'),
			('Start Date', start_date),
			('Announcement Date', row[1].value),
			('End Date', row[2].value),
			('Formated Start Date', int(start_date.strftime('%Y%m%d'))),
			('Formated Announcement Date', int(row[1].value.date().strftime('%Y%m%d'))),
			('Formated End Date', int(row[2].value.strftime('%Y%m%d'))),
		]

		#create the file from source wb 
		self.wb_generator.new_company_workbook(3, self.acquirer_path)
		#load the newly created wb
		test_wb_file = os.listdir(self.acquirer_path)[0]
		path = os.path.join(self.acquirer_path, test_wb_file)
		test_wb = load_workbook(path)
		ws = test_wb['Options Chain']
		for row in range(9):
			for col in range(2):
				self.assertEqual(ws.cell(column=col+1, row=row+1).value,data[row][col])


	def test_get_company_options_tickers(self):
		# def get_company_options_tickers(self,reference_sheet, start_date, 
		#	announcement_date, row, start_column, interval, ticker_cell, type_cell):
		#test the cells where the add_BDS_OPT_CHAIN function gets added
		wb = Workbook()
		st_date = dt.date.today()-dt.timedelta(days=360)
		ann_date = dt.date.today()
		row = 1
		col = 1
		interval = 30
		ticker_cell = 'B2'
		type_cell = 'B3'

		sheet = wb.active
		self.wb_generator.get_company_options_tickers(sheet, st_date,
			ann_date, row, col, interval, ticker_cell, type_cell)

		for i in range(1, sheet.max_column+1, 2):
			self.assertIn( '=BDS(', sheet.cell(row=row, column=i).value)

	def test_create_company_workbooks(self):
		self.assertEqual(len(os.listdir(self.target_path)), 0)
		self.assertEqual(len(os.listdir(self.acquirer_path)), 0)
		self.wb_generator.create_company_workbooks()
		self.assertGreater(len(os.listdir(self.target_path)), 0)
		self.assertGreater(len(os.listdir(self.target_path)), 0)


if __name__ == "__main__":
	unittest.main()